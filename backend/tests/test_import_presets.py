"""取り込み設定（プリセット）とブック一括取り込み.

The point of the feature: a workbook with many worksheets should only ever be set
up once. These tests pin the parts that make the second run cheap —

- finishing an import saves the settings, keyed by the SOURCE worksheet name;
- a re-run matches those settings and UPDATES the same sheet instead of making a
  second one (upsert by ID);
- a worksheet with no saved setting is skipped, never silently imported;
- a failure anywhere rolls the whole book back.
"""
from __future__ import annotations

import io
import json

from openpyxl import Workbook

from .conftest import make_sheet

_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _book() -> io.BytesIO:
    """Two data worksheets plus a cover sheet, the shape these books really have."""
    wb = Workbook()
    cover = wb.active
    cover.title = "表紙"
    cover.append(["社外秘"])

    a = wb.create_sheet("設計スケジュール")
    a.append(["ID", "件名", "区分"])
    a.append(["A-1", "基本設計", "開発"])
    a.append(["A-2", "詳細設計", "開発"])

    b = wb.create_sheet("製造スケジュール")
    b.append(["ID", "件名", "区分"])
    b.append(["B-1", "実装", "開発"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _presets(client) -> dict[str, dict]:
    r = client.get("/api/import/presets")
    assert r.status_code == 200, r.text
    return {p["worksheet_name"]: p for p in r.json()}


def test_wizard_import_is_remembered_as_a_preset(auth_client):
    """The 既存シート wizard's settings are saved under the worksheet's name, so the
    next round can replay them without asking anything."""
    sid = make_sheet(auth_client, "設計")
    auth_client.post(f"/api/sheets/{sid}/columns", json={"name": "件名", "type": "text"})

    r = auth_client.post(
        "/api/import/presets",
        json={
            "worksheet_name": "設計スケジュール",
            "workbook_name": "計画.xlsx",
            "target_sheet_id": sid,
            "header_row": 1,
            "id_column": 0,
            "mapping": [{"index": 1, "name": "件名", "type": ""}],
        },
    )
    assert r.status_code == 200, r.text

    saved = _presets(auth_client)["設計スケジュール"]
    assert saved["target_sheet_id"] == sid
    assert saved["mapping"] == [{"index": 1, "name": "件名", "type": ""}]

    # Saving again for the same worksheet REPLACES it (one setting per worksheet).
    auth_client.post(
        "/api/import/presets",
        json={"worksheet_name": "設計スケジュール", "target_sheet_id": sid, "header_row": 3},
    )
    again = _presets(auth_client)
    assert len(again) == 1
    assert again["設計スケジュール"]["header_row"] == 3


def test_inspect_workbook_skips_unknown_worksheets(auth_client):
    """Dropping a book must not silently create a sheet per worksheet — anything
    without a saved setting defaults to 取り込まない."""
    r = auth_client.post(
        "/api/import/workbook/inspect", files={"file": ("計画.xlsx", _book(), _MEDIA)}
    )
    assert r.status_code == 200, r.text
    by_name = {w["worksheet"]: w for w in r.json()["worksheets"]}
    assert set(by_name) == {"表紙", "設計スケジュール", "製造スケジュール"}
    assert all(w["action"] == "skip" for w in by_name.values())
    assert all(w["preset_id"] is None for w in by_name.values())


def test_workbook_run_creates_sheets_then_updates_them(auth_client):
    """The whole flow: first run creates the sheets, and because the settings are
    written back pointing at those sheets, the SECOND run of the same book updates
    them in place instead of duplicating."""
    plan = [
        {"worksheet": "設計スケジュール", "action": "new", "target_sheet_name": "設計"},
        {"worksheet": "製造スケジュール", "action": "new", "target_sheet_name": "製造"},
    ]
    r = auth_client.post(
        "/api/import/workbook",
        files={"file": ("計画.xlsx", _book(), _MEDIA)},
        data={"plan": json.dumps(plan)},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 3  # 2 + 1 rows
    assert body["updated"] == 0
    assert {x["worksheet"] for x in body["results"]} == {"設計スケジュール", "製造スケジュール"}
    first_ids = {x["worksheet"]: x["sheet_id"] for x in body["results"]}

    # 表紙 was not in the plan, so it stayed out of it.
    assert len(auth_client.get("/api/sheets").json()) == 2

    # The settings now point at the sheets that were created.
    saved = _presets(auth_client)
    assert saved["設計スケジュール"]["target_sheet_id"] == first_ids["設計スケジュール"]

    # Second run: no plan at all — the saved settings alone drive it.
    r2 = auth_client.post(
        "/api/import/workbook", files={"file": ("計画.xlsx", _book(), _MEDIA)}
    )
    assert r2.status_code == 200, r2.text
    body2 = r2.json()
    assert body2["created"] == 0
    assert body2["updated"] == 3  # every row matched by ID
    assert len(auth_client.get("/api/sheets").json()) == 2  # no duplicates
    assert {x["sheet_id"] for x in body2["results"]} == set(first_ids.values())


def test_inspect_workbook_reports_counts_for_saved_worksheets(auth_client):
    """Once a setting exists the dry-run says what the re-run would do, so the
    whole book can be checked before writing anything."""
    auth_client.post(
        "/api/import/workbook",
        files={"file": ("計画.xlsx", _book(), _MEDIA)},
        data={"plan": json.dumps([{"worksheet": "設計スケジュール", "action": "new"}])},
    )

    r = auth_client.post(
        "/api/import/workbook/inspect", files={"file": ("計画.xlsx", _book(), _MEDIA)}
    )
    assert r.status_code == 200, r.text
    by_name = {w["worksheet"]: w for w in r.json()["worksheets"]}
    designed = by_name["設計スケジュール"]
    assert designed["action"] == "existing"
    assert designed["preset_id"] is not None
    assert designed["updated_rows"] == 2 and designed["new_rows"] == 0
    assert designed["error"] is None
    # Untouched worksheets are still opt-in.
    assert by_name["製造スケジュール"]["action"] == "skip"


def test_workbook_run_rolls_everything_back_on_failure(auth_client):
    """All-or-nothing: with a dozen worksheets, a half-written book is worse than
    nothing, so one bad worksheet undoes the ones that already succeeded."""
    plan = [
        {"worksheet": "設計スケジュール", "action": "new", "target_sheet_name": "設計"},
        # 表紙 has a single cell and no usable 見出し行 for the mapping asked of it.
        {"worksheet": "表紙", "action": "existing", "target_sheet_id": 999999},
    ]
    r = auth_client.post(
        "/api/import/workbook",
        files={"file": ("計画.xlsx", _book(), _MEDIA)},
        data={"plan": json.dumps(plan)},
    )
    # 表紙 falls back to 新規 (the target does not exist) and then fails on its
    # empty 見出し行 — either way nothing may survive.
    if r.status_code != 200:
        assert auth_client.get("/api/sheets").json() == []
        assert _presets(auth_client) == {}


def _book_with_junk_footer() -> io.BytesIO:
    """The realistic shape: data, a blank line, then a 合計 row and a note that
    must not become tasks. Data is rows 2-4; the junk starts at row 6."""
    wb = Workbook()
    ws = wb.active
    ws.title = "設計スケジュール"
    ws.append(["ID", "件名", "区分"])      # row 1
    ws.append(["A-1", "基本設計", "開発"])  # row 2
    ws.append(["A-2", "詳細設計", "開発"])  # row 3
    ws.append(["A-3", "設計レビュー", "開発"])  # row 4
    ws.append([])                          # row 5
    ws.append(["合計", "3件", ""])          # row 6
    ws.append(["※ 2026年度分", "", ""])     # row 7
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_last_row_cuts_the_footer_off(auth_client):
    """「これ以降を取り込まない」: without it the 合計行 and the note become tasks."""
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("計画.xlsx", _book_with_junk_footer(), _MEDIA)},
        data={"name": "切らない", "has_week_grid": "true"},
    )
    assert r.status_code == 201, r.text
    assert r.json()["created"] == 5  # 3 tasks + 合計 + 注記

    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("計画.xlsx", _book_with_junk_footer(), _MEDIA)},
        data={"name": "切る", "has_week_grid": "true", "last_row": "4"},
    )
    assert r.status_code == 201, r.text
    assert r.json()["created"] == 3

    detail = auth_client.get(f"/api/sheets/{r.json()['sheet_id']}").json()
    assert {row["key_value"] for row in detail["rows"]} == {"A-1", "A-2", "A-3"}


def test_inspect_reports_the_cut_and_shows_the_tail(auth_client):
    """The picker needs the BOTTOM rows — the head preview stops at row 30, but the
    rows to cut are at the end."""
    r = auth_client.post(
        "/api/sheets/import.xlsx/inspect",
        files={"file": ("計画.xlsx", _book_with_junk_footer(), _MEDIA)},
        data={"has_week_grid": "true", "last_row": "4"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["last_row"] == 4
    assert body["sheet_last_row"] == 7
    assert body["total_rows"] == 3       # after the cut
    assert body["available_rows"] == 5   # before it
    # The tail preview reaches the junk and never above the 見出し行.
    tail_rows = [t["row"] for t in body["tail_preview"]]
    assert 6 in tail_rows and 7 in tail_rows
    assert min(tail_rows) > body["header_row"]


def _long_book(rows: int = 148) -> io.BytesIO:
    """Long enough that the cut can be well above the last preview window."""
    wb = Workbook()
    ws = wb.active
    ws.title = "設計スケジュール"
    ws.append(["ID", "件名"])
    for i in range(1, rows):
        ws.append([f"A-{i}", f"作業{i}"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _tail(client, **data) -> list[int]:
    r = client.post(
        "/api/sheets/import.xlsx/inspect",
        files={"file": ("長い.xlsx", _long_book(), _MEDIA)},
        data={"has_week_grid": "true", **data},
    )
    assert r.status_code == 200, r.text
    return [t["row"] for t in r.json()["tail_preview"]]


def test_tail_window_can_reach_rows_far_above_the_bottom(auth_client):
    """The cut is often nowhere near the last few rows, so the tail window must be
    able to move up — otherwise a row 80 cut on a 148-row sheet is unclickable."""
    auto = _tail(auth_client)
    assert auto[-1] == 148 and len(auto) == 15  # 末尾15行が既定

    # 「さらに上を表示」 walks the window up to an explicit row.
    walked = _tail(auth_client, tail_from="100")
    assert walked[0] == 100 and walked[-1] == 148

    # A row TYPED into 最終行 drags the window up on its own, with context above
    # it — so the cut can always be eyeballed instead of set blind.
    typed = _tail(auth_client, last_row="80")
    assert typed[0] < 80 < typed[-1]

    # Never above the 見出し行, however far up it is pushed.
    assert min(_tail(auth_client, last_row="1", tail_from="1")) > 1


def test_last_row_is_remembered_and_replayed(auth_client):
    """The cut is a property of the source sheet, so it rides along with the rest
    of the setting — the next run must not re-import the footer."""
    plan = [{"worksheet": "設計スケジュール", "action": "new", "target_sheet_name": "設計", "last_row": 4}]
    r = auth_client.post(
        "/api/import/workbook",
        files={"file": ("計画.xlsx", _book_with_junk_footer(), _MEDIA)},
        data={"plan": json.dumps(plan)},
    )
    assert r.status_code == 200, r.text
    assert r.json()["created"] == 3
    assert _presets(auth_client)["設計スケジュール"]["last_row"] == 4

    # Second run with NO plan — the saved cut alone has to hold.
    r2 = auth_client.post(
        "/api/import/workbook", files={"file": ("計画.xlsx", _book_with_junk_footer(), _MEDIA)}
    )
    assert r2.status_code == 200, r2.text
    assert r2.json() == {**r2.json(), "created": 0, "updated": 3}

    entry = {
        w["worksheet"]: w
        for w in auth_client.post(
            "/api/import/workbook/inspect",
            files={"file": ("計画.xlsx", _book_with_junk_footer(), _MEDIA)},
        ).json()["worksheets"]
    }["設計スケジュール"]
    assert entry["last_row"] == 4
    assert entry["total_rows"] == 3 and entry["available_rows"] == 5
    assert any("除外" in w for w in entry["warnings"])


def test_lookup_columns_are_never_overwritten_by_a_reimport(auth_client):
    """参照(LOOKUP)列は計算値なので、取り込みは触らない。取り込んだ後で列を LOOKUP に
    変えても、再取り込みでその列に Excel の値が書き戻ることはない。"""
    sid = make_sheet(auth_client, "設計")
    r = auth_client.post(f"/api/sheets/{sid}/columns", json={"name": "件名", "type": "text"})
    title = r.json()["id"]
    r = auth_client.post(f"/api/sheets/{sid}/columns", json={"name": "区分", "type": "text"})
    kind = r.json()["id"]

    book = lambda: _xlsx_rows([["ID", "件名", "区分"], ["A-1", "基本設計", "開発"]])
    assert auth_client.post(
        f"/api/sheets/{sid}/import.xlsx", files={"file": ("x.xlsx", book(), _MEDIA)}
    ).status_code == 200
    rows = auth_client.get(f"/api/sheets/{sid}").json()["rows"]
    assert rows[0]["data"][str(kind)] == "開発"

    # 区分 を参照(LOOKUP)列に変更 → 以後この列は計算値。
    assert auth_client.patch(
        f"/api/columns/{kind}",
        json={"type": "lookup", "config": {"target_sheet_id": sid,
                                           "match_key_column_id": title,
                                           "return_column_id": title}},
    ).status_code == 200

    # 同じ Excel を取り込み直しても、LOOKUP 列は無視される（件名は更新される）。
    book2 = _xlsx_rows([["ID", "件名", "区分"], ["A-1", "基本設計v2", "別の値"]])
    assert auth_client.post(
        f"/api/sheets/{sid}/import.xlsx", files={"file": ("x.xlsx", book2, _MEDIA)}
    ).status_code == 200
    data = auth_client.get(f"/api/sheets/{sid}").json()["rows"][0]["data"]
    assert data[str(title)] == "基本設計v2"      # 通常列は更新される
    assert data.get(str(kind)) == "開発"          # LOOKUP 列は書き換わらない

    # ドライランも「取り込む列」に数えず、理由を出す。
    r = auth_client.post(
        f"/api/sheets/{sid}/import.xlsx/inspect",
        files={"file": ("x.xlsx", book(), _MEDIA)},
        data={"columns": json.dumps([{"index": 1, "name": "件名"}, {"index": 2, "name": "区分"}])},
    )
    by_header = {c["header"]: c for c in r.json()["columns"]}
    assert by_header["区分"]["target"] == ""          # 取り込み先から外れる
    assert by_header["区分"]["lost_reason"] == "lookup"
    assert by_header["件名"]["target"] == "件名"


def test_a_renamed_column_is_reported_not_silently_dropped(auth_client):
    """保存済みマッピングは列「名」を持つので、改名すると取り込みでは無視される。
    無視すること自体は安全だが、件数だけ黙って減るのは困るので理由を出す。"""
    sid = make_sheet(auth_client, "設計")
    r = auth_client.post(f"/api/sheets/{sid}/columns", json={"name": "件名", "type": "text"})
    title = r.json()["id"]
    assert auth_client.patch(f"/api/columns/{title}", json={"name": "タイトル"}).status_code == 200

    r = auth_client.post(
        f"/api/sheets/{sid}/import.xlsx/inspect",
        files={"file": ("x.xlsx", _xlsx_rows([["ID", "件名"], ["A-1", "基本設計"]]), _MEDIA)},
        data={"columns": json.dumps([{"index": 1, "name": "件名"}])},
    )
    col = {c["header"]: c for c in r.json()["columns"]}["件名"]
    assert col["target"] == "" and col["lost_reason"] == "missing"
    assert col["lost_target"] == "件名"


def _xlsx_rows(rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_preset_can_be_deleted(auth_client):
    sid = make_sheet(auth_client, "設計")
    auth_client.post(
        "/api/import/presets",
        json={"worksheet_name": "設計スケジュール", "target_sheet_id": sid},
    )
    pid = _presets(auth_client)["設計スケジュール"]["id"]
    assert auth_client.delete(f"/api/import/presets/{pid}").status_code == 204
    assert _presets(auth_client) == {}


def test_preset_survives_its_sheet_being_deleted(auth_client):
    """A preset whose sheet is gone degrades to 新規作成 rather than disappearing —
    the saved intent is the expensive part, the sheet is not."""
    sid = make_sheet(auth_client, "設計")
    auth_client.post(
        "/api/import/presets",
        json={"worksheet_name": "設計スケジュール", "target_sheet_id": sid, "header_row": 2},
    )
    assert auth_client.delete(f"/api/sheets/{sid}").status_code in (200, 204)

    saved = _presets(auth_client)["設計スケジュール"]
    assert saved["target_sheet_id"] is None
    assert saved["header_row"] == 2

    r = auth_client.post(
        "/api/import/workbook/inspect", files={"file": ("計画.xlsx", _book(), _MEDIA)}
    )
    entry = {w["worksheet"]: w for w in r.json()["worksheets"]}["設計スケジュール"]
    assert entry["action"] == "new"
    assert any("見つかりません" in w for w in entry["warnings"])
