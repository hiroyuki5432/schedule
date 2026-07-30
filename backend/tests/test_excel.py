"""Excel export/import round-trip (upsert by ID, attributes + weekly effort)."""
from __future__ import annotations

import io
import json

from openpyxl import Workbook, load_workbook

from .conftest import make_sheet

_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# A phase → ◇ → phase template (既定マイルストン).
TEMPLATE = [
    {"name": "設計", "kind": "phase", "weight": 1},
    {"name": "レビュー", "kind": "milestone"},
    {"name": "実装", "kind": "phase", "weight": 1},
]


def _add_text_column(client, sheet_id: int, name: str) -> int:
    r = client.post(f"/api/sheets/{sheet_id}/columns", json={"name": name, "type": "text"})
    assert r.status_code in (200, 201), r.text
    return r.json()["id"]


def _set_template(client, sheet_id: int, items: list) -> None:
    r = client.patch(f"/api/sheets/{sheet_id}", json={"settings": {"default_milestones": items}})
    assert r.status_code == 200, r.text


def _sched_col_ids(client, sheet_id: int) -> tuple[str, str]:
    """GET the sheet (creates 開始日/完了日 columns) and return their (start, end) ids."""
    detail = client.get(f"/api/sheets/{sheet_id}").json()
    start = end = None
    for c in detail["columns"]:
        role = (c.get("config") or {}).get("sched_role")
        if role == "start":
            start = str(c["id"])
        elif role == "end":
            end = str(c["id"])
    assert start and end
    return start, end


def test_export_xlsx_has_header_and_rows(auth_client):
    sid = make_sheet(auth_client, "X")
    col = _add_text_column(auth_client, sid, "件名")
    auth_client.post(f"/api/sheets/{sid}/rows", json={"key_value": "A-1", "data": {str(col): "設計"}})

    r = auth_client.get(f"/api/sheets/{sid}/export.xlsx")
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[0] == "ID"
    assert "件名" in header
    values = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert any(v[0] == "A-1" for v in values)


def test_import_upserts_by_id(auth_client):
    sid = make_sheet(auth_client, "Y")
    col = _add_text_column(auth_client, sid, "件名")
    auth_client.post(f"/api/sheets/{sid}/rows", json={"key_value": "A-1", "data": {str(col): "旧"}})

    # Build an xlsx: update A-1, add new B-2.
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "件名"])
    ws.append(["A-1", "新"])
    ws.append(["B-2", "追加"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        f"/api/sheets/{sid}/import.xlsx",
        files={"file": ("in.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    assert r.json() == {"created": 1, "updated": 1}

    rows = auth_client.get(f"/api/sheets/{sid}/rows").json()
    by_key = {row["key_value"]: row for row in rows}
    assert by_key["A-1"]["data"][str(col)] == "新"  # updated, not duplicated
    assert by_key["B-2"]["data"][str(col)] == "追加"  # newly created
    assert len(rows) == 2


def test_import_ignores_lookup_columns(auth_client):
    sid = make_sheet(auth_client, "L")
    col = _add_text_column(auth_client, sid, "件名")
    lk = auth_client.post(
        f"/api/sheets/{sid}/columns",
        json={"name": "参照値", "type": "lookup", "config": {}},
    )
    assert lk.status_code in (200, 201), lk.text
    lk_id = lk.json()["id"]

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "件名", "参照値"])
    ws.append(["A-1", "設計", "勝手な値"])  # lookup cell should be ignored
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        f"/api/sheets/{sid}/import.xlsx",
        files={"file": ("in.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    rows = auth_client.get(f"/api/sheets/{sid}/rows").json()
    data = rows[0]["data"]
    assert data[str(col)] == "設計"
    assert str(lk_id) not in data  # lookup value never written


def test_export_has_sched_and_template_milestone_columns(auth_client):
    sid = make_sheet(auth_client, "M")
    _add_text_column(auth_client, sid, "件名")
    _set_template(auth_client, sid, TEMPLATE)
    start_id, end_id = _sched_col_ids(auth_client, sid)

    r = auth_client.post(
        f"/api/sheets/{sid}/rows",
        json={"key_value": "A-1", "data": {start_id: "2026-01-01", end_id: "2026-03-01"}},
    )
    rid = r.json()["id"]
    auth_client.put(
        f"/api/rows/{rid}/milestones",
        json=[
            {"name": "設計", "kind": "phase", "boundary_date": "2026-01-01", "order": 0},
            {"name": "レビュー", "kind": "milestone", "boundary_date": "2026-02-01",
             "order": 1, "done": True, "actual_date": "2026-02-05"},
            {"name": "実装", "kind": "phase", "boundary_date": "2026-02-01", "order": 2},
        ],
    )

    ws = load_workbook(io.BytesIO(auth_client.get(f"/api/sheets/{sid}/export.xlsx").content)).active
    header = [c.value for c in ws[1]]
    for h in ("開始日", "完了日", "進捗(%)", "先行タスク(ID)", "レビュー（予定）", "レビュー（実績）"):
        assert h in header, h
    a1 = next([c.value for c in row] for row in ws.iter_rows(min_row=2) if row[0].value == "A-1")
    assert a1[header.index("開始日")] == "2026-01-01"
    assert a1[header.index("レビュー（予定）")] == "2026-02-01"
    assert a1[header.index("レビュー（実績）")] == "2026-02-05"


def test_export_clear_import_round_trip(auth_client):
    """Export → clear the sheet → re-import fully restores rows, span, milestones,
    進捗 and 先行タスク (deps round-trip by key_value, resolved after import)."""
    sid = make_sheet(auth_client, "RT")
    col = _add_text_column(auth_client, sid, "件名")
    _set_template(auth_client, sid, TEMPLATE)
    start_id, end_id = _sched_col_ids(auth_client, sid)

    auth_client.post(f"/api/sheets/{sid}/rows", json={"key_value": "T-0", "data": {str(col): "前工程"}})
    r = auth_client.post(
        f"/api/sheets/{sid}/rows",
        json={"key_value": "T-1", "data": {str(col): "作業", start_id: "2026-01-01", end_id: "2026-03-01"}},
    )
    body = r.json()
    rid = body["id"]
    t0_id = next(
        x["id"] for x in auth_client.get(f"/api/sheets/{sid}/rows").json() if x["key_value"] == "T-0"
    )
    auth_client.patch(
        f"/api/rows/{rid}",
        json={"data": body["data"], "version": body["version"], "progress": 40, "depends_on": [int(t0_id)]},
    )
    auth_client.put(
        f"/api/rows/{rid}/milestones",
        json=[
            {"name": "設計", "kind": "phase", "boundary_date": "2026-01-01", "order": 0},
            {"name": "レビュー", "kind": "milestone", "boundary_date": "2026-02-01", "order": 1},
            {"name": "実装", "kind": "phase", "boundary_date": "2026-02-01", "order": 2},
        ],
    )

    blob = auth_client.get(f"/api/sheets/{sid}/export.xlsx").content
    assert auth_client.delete(f"/api/sheets/{sid}/rows").status_code == 200
    r = auth_client.post(f"/api/sheets/{sid}/import.xlsx", files={"file": ("rt.xlsx", io.BytesIO(blob), _MEDIA)})
    assert r.status_code == 200, r.text

    rows = {x["key_value"]: x for x in auth_client.get(f"/api/sheets/{sid}/rows").json()}
    t1 = rows["T-1"]
    assert t1["data"][start_id] == "2026-01-01"
    assert t1["progress"] == 40
    assert t1["depends_on"] == [int(rows["T-0"]["id"])]
    ms = sorted(auth_client.get(f"/api/rows/{t1['id']}/milestones").json(), key=lambda m: m["order"])
    assert [(m["name"], m["kind"], m["boundary_date"]) for m in ms] == [
        ("設計", "phase", "2026-01-01"),
        ("レビュー", "milestone", "2026-02-01"),
        ("実装", "phase", "2026-02-01"),  # phase boundary reconstructed from ◇
    ]


def test_export_has_week_columns_without_effort(auth_client):
    """A sheet with rows but no effort yet must still export fillable week columns
    spanning 開始日〜完了日, so 工数 can be entered and re-imported (要望: 工数Excel取込)."""
    sid = make_sheet(auth_client, "WK")
    _add_text_column(auth_client, sid, "件名")
    start_id, end_id = _sched_col_ids(auth_client, sid)
    auth_client.post(
        f"/api/sheets/{sid}/rows",
        json={"key_value": "E-1", "data": {start_id: "2026-02-02", end_id: "2026-02-20"}},
    )

    ws = load_workbook(io.BytesIO(auth_client.get(f"/api/sheets/{sid}/export.xlsx").content)).active
    header = [c.value for c in ws[1]]
    # Week columns are ISO date headers (Mondays). At least the row's span weeks.
    week_headers = [h for h in header if isinstance(h, str) and h.startswith("2026-02")]
    assert week_headers, header

    # Fill the first week column with hours, then import → effort is created.
    wb = Workbook()
    out = wb.active
    out.append(["ID", "件名", week_headers[0]])
    out.append(["E-1", "作業", 8])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = auth_client.post(
        f"/api/sheets/{sid}/import.xlsx", files={"file": ("wk.xlsx", buf, _MEDIA)}
    )
    assert r.status_code == 200, r.text
    effort = auth_client.get(f"/api/sheets/{sid}/effort").json()
    assert len(effort) == 1
    assert float(effort[0]["actual_hours"] or 0) + float(effort[0]["planned_hours"] or 0) == 8.0


def _xlsx(rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_import_creates_new_sheet_with_inferred_columns(auth_client):
    """POST /api/sheets/import.xlsx builds a sheet from the header row and guesses
    each column's type from its values (要望: シートもexcelから取り込める)."""
    buf = _xlsx(
        [
            ["ID", "件名", "担当", "区分", "開始予定", "見積"],
            ["A-1", "設計する", "Admin", "開発", "2026-04-01", 12],
            ["A-2", "実装する", "Admin", "開発", "2026-04-08", 30],
            ["A-3", "確認する", "Admin", "検証", "2026-04-15", 4],
        ]
    )
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("plan.xlsx", buf, _MEDIA)},
        data={"name": "取込テスト", "has_week_grid": "true"},
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["created"] == 3
    sid = body["sheet_id"]

    detail = auth_client.get(f"/api/sheets/{sid}").json()
    assert detail["sheet"]["name"] == "取込テスト"
    by_name = {c["name"]: c for c in detail["columns"]}
    assert by_name["件名"]["type"] == "text"
    assert by_name["担当"]["type"] == "member"  # matches an org member's name
    assert by_name["区分"]["type"] == "dropdown"  # few repeated labels
    assert sorted(o["value"] for o in by_name["区分"]["config"]["options"]) == ["検証", "開発"]
    assert by_name["開始予定"]["type"] == "date"
    assert by_name["見積"]["type"] == "number"

    rows = {x["key_value"]: x for x in detail["rows"]}
    assert rows["A-1"]["data"][str(by_name["件名"]["id"])] == "設計する"
    assert rows["A-1"]["data"][str(by_name["担当"]["id"])] == auth_client.org_admin["admin_id"]
    assert rows["A-3"]["data"][str(by_name["見積"]["id"])] == 4


def test_import_new_sheet_keeps_week_columns_out_of_attributes(auth_client):
    """ISO-date headers stay week columns (工数) instead of becoming attributes."""
    buf = _xlsx([["ID", "件名", "2099-01-05"], ["W-1", "作業", 9]])
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("wk.xlsx", buf, _MEDIA)},
        data={"has_week_grid": "true"},
    )
    assert r.status_code == 201, r.text
    sid = r.json()["sheet_id"]

    names = {c["name"] for c in auth_client.get(f"/api/sheets/{sid}").json()["columns"]}
    assert "2099-01-05" not in names
    assert {"件名", "開始日", "完了日"} <= names

    effort = auth_client.get(f"/api/sheets/{sid}/effort").json()
    assert len(effort) == 1
    assert float(effort[0]["planned_hours"]) == 9.0


def test_import_new_sheet_defaults_name_to_worksheet(auth_client):
    wb = Workbook()
    ws = wb.active
    ws.title = "作業一覧"
    ws.append(["ID", "件名"])
    ws.append(["N-1", "メモ"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("x.xlsx", buf, _MEDIA)},
        data={"has_week_grid": "false"},
    )
    assert r.status_code == 201, r.text
    sid = r.json()["sheet_id"]
    detail = auth_client.get(f"/api/sheets/{sid}").json()
    assert detail["sheet"]["name"] == "作業一覧"
    assert detail["sheet"]["has_week_grid"] is False


def _multi_sheet_xlsx() -> io.BytesIO:
    """A messy but realistic book: a cover sheet + a data sheet whose header is on
    row 3 (title line, blank line), with the ID in column B."""
    wb = Workbook()
    cover = wb.active
    cover.title = "表紙"
    cover.append(["社外秘"])
    ws = wb.create_sheet("作業計画")
    ws.append(["2026年度 作業計画"])
    ws.append([])
    ws.append(["No", "タスクID", "件名", "担当", "区分", "開始予定", "見積"])
    ws.append([1, "A-1", "設計する", "Admin", "開発", "2026-04-01", 12])
    ws.append([2, "A-2", "実装する", "Admin", "開発", "2026-04-08", 30])
    ws.append([3, "A-3", "確認する", "Admin", "検証", "2026-04-15", 4])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_inspect_lists_worksheets_and_guesses_header_row(auth_client):
    """取り込みウィザード step 1/2: worksheets, the guessed 見出し行 and a raw preview."""
    r = auth_client.post(
        "/api/sheets/import.xlsx/inspect",
        files={"file": ("plan.xlsx", _multi_sheet_xlsx(), _MEDIA)},
        data={"sheet_name": "作業計画", "id_column": "1", "has_week_grid": "true"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert [w["name"] for w in body["worksheets"]] == ["表紙", "作業計画"]
    assert body["sheet_name"] == "作業計画"
    assert body["suggested_header_row"] == 3  # skips the title line + blank row
    assert body["header_row"] == 3
    assert body["total_rows"] == 3
    assert body["preview"][2]["cells"][:3] == ["No", "タスクID", "件名"]

    by_header = {c["header"]: c for c in body["columns"]}
    assert by_header["タスクID"]["selected"] is False  # it's the ID column
    assert by_header["担当"]["type"] == "member"
    assert by_header["区分"]["type"] == "dropdown"
    assert sorted(by_header["区分"]["options"]) == ["検証", "開発"]
    assert by_header["開始予定"]["type"] == "date"
    assert by_header["見積"]["type"] == "number"
    assert by_header["件名"]["samples"] == ["設計する", "実装する", "確認する"]
    assert by_header["件名"]["filled"] == 3


def test_inspect_reports_values_that_would_not_convert(auth_client):
    """The preview step warns before anything is written: cells that don't fit the
    chosen type are counted (and sampled) instead of silently vanishing."""
    buf = _xlsx([["ID", "期日"], ["A-1", "2026-04-01"], ["A-2", "未定"]])
    r = auth_client.post(
        "/api/sheets/import.xlsx/inspect",
        files={"file": ("d.xlsx", buf, _MEDIA)},
        data={"columns": '[{"index": 1, "name": "期日", "type": "date"}]'},
    )
    assert r.status_code == 200, r.text
    col = next(c for c in r.json()["columns"] if c["header"] == "期日")
    assert col["selected"] is True
    assert col["invalid"] == 1
    assert col["invalid_samples"] == ["未定"]


def test_inspect_flags_duplicate_and_blank_ids(auth_client):
    buf = _xlsx([["ID", "件名"], ["A-1", "あ"], ["A-1", "い"], [None, "う"]])
    r = auth_client.post(
        "/api/sheets/import.xlsx/inspect",
        files={"file": ("dup.xlsx", buf, _MEDIA)},
    )
    body = r.json()
    assert body["duplicate_ids"] == 1
    assert body["blank_ids"] == 1


def test_import_honors_worksheet_header_row_and_column_choice(auth_client):
    """Only the chosen worksheet / 見出し行 / ID列 / 列 are taken, with the user's
    own column names and types (要望: 確認しながら取り込む)."""
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("plan.xlsx", _multi_sheet_xlsx(), _MEDIA)},
        data={
            "name": "選択取込",
            "has_week_grid": "true",
            "sheet_name": "作業計画",
            "header_row": "3",
            "id_column": "1",
            "columns": json.dumps(
                [
                    {"index": 2, "name": "作業名", "type": "text"},
                    {"index": 5, "name": "開始日", "type": "date"},
                    {"index": 6, "name": "見積", "type": "text"},  # forced to text
                ]
            ),
        },
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["created"] == 3
    sid = body["sheet_id"]

    detail = auth_client.get(f"/api/sheets/{sid}").json()
    by_name = {c["name"]: c for c in detail["columns"]}
    assert "No" not in by_name and "担当" not in by_name and "区分" not in by_name
    assert by_name["作業名"]["type"] == "text"
    assert by_name["見積"]["type"] == "text"  # user's override beats the number guess
    # 開始日 binds to the sheet's existing schedule column instead of duplicating it.
    assert by_name["開始日"]["config"].get("sched_role") == "start"
    assert body["columns"] == 2

    rows = {x["key_value"]: x for x in detail["rows"]}
    assert set(rows) == {"A-1", "A-2", "A-3"}
    assert rows["A-1"]["data"][str(by_name["作業名"]["id"])] == "設計する"
    assert rows["A-1"]["data"][str(by_name["開始日"]["id"])] == "2026-04-01"
    assert rows["A-3"]["data"][str(by_name["見積"]["id"])] == "4"


def test_import_without_id_column_auto_numbers_rows(auth_client):
    buf = _xlsx([["件名", "区分"], ["設計", "開発"], ["実装", "開発"]])
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("noid.xlsx", buf, _MEDIA)},
        data={
            "has_week_grid": "false",
            "id_column": "-1",
            "columns": json.dumps(
                [{"index": 0, "name": "件名", "type": "text"}, {"index": 1, "name": "区分", "type": "text"}]
            ),
        },
    )
    assert r.status_code == 201, r.text
    detail = auth_client.get(f"/api/sheets/{r.json()['sheet_id']}").json()
    by_name = {c["name"]: c for c in detail["columns"]}
    keys = sorted(x["key_value"] for x in detail["rows"])
    assert keys == ["001", "002"]  # numbering rule filled the missing IDs
    assert detail["rows"][0]["data"][str(by_name["件名"]["id"])] == "設計"


def test_import_weekly_effort_future_planned(auth_client):
    sid = make_sheet(auth_client, "Z")  # week-grid sheet
    _add_text_column(auth_client, sid, "件名")

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "件名", "2099-01-05"])  # far-future week → planned
    ws.append(["W-1", "作業", 12])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        f"/api/sheets/{sid}/import.xlsx",
        files={"file": ("in.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text

    effort = auth_client.get(f"/api/sheets/{sid}/effort").json()
    assert len(effort) == 1
    assert float(effort[0]["planned_hours"]) == 12.0
    assert effort[0]["actual_hours"] is None
