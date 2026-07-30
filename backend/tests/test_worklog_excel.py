"""Worklog (みんなの入力一覧) Excel export/import."""
from __future__ import annotations

import io
import json

from openpyxl import Workbook, load_workbook


def test_export_worklog_xlsx(auth_client):
    auth_client.post(
        "/api/worklog",
        json={"work_date": "2026-06-01", "cat1": "開発", "memo": "設計", "hours": 3},
    )
    r = auth_client.get("/api/worklog/export.xlsx?from=2026-06-01&to=2026-06-01")
    assert r.status_code == 200, r.text
    ws = load_workbook(io.BytesIO(r.content)).active
    header = [c.value for c in ws[1]]
    assert header == ["日付", "ユーザー", "タスクID", "大分類", "中分類", "メモ", "時間"]
    body = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert any(row[5] == "設計" and float(row[6]) == 3.0 for row in body)


def test_import_worklog_adds_logs(auth_client):
    admin_name = "Admin"  # from conftest org_admin fixture
    wb = Workbook()
    ws = wb.active
    ws.append(["日付", "ユーザー", "タスクID", "大分類", "中分類", "メモ", "時間"])
    ws.append(["2026-06-02", admin_name, "", "会議", "", "定例", 1.5])
    ws.append(["2026-06-02", "存在しない人", "", "x", "", "", 9])  # skipped
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        "/api/worklog/import.xlsx",
        files={"file": ("wl.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    assert r.json() == {"created": 1, "skipped": 1, "duplicates": 0}

    logs = auth_client.get("/api/worklog?from=2026-06-02&to=2026-06-02").json()
    assert any(l["memo"] == "定例" and float(l["hours"]) == 1.5 for l in logs)


def test_import_worklog_dedupes_identical_rows(auth_client):
    """Re-importing the same row is skipped as a duplicate (no double-count)."""
    def _file():
        wb = Workbook()
        ws = wb.active
        ws.append(["日付", "ユーザー", "タスクID", "大分類", "中分類", "メモ", "時間"])
        ws.append(["2026-06-03", "Admin", "", "開発", "", "実装", 2])
        ws.append(["2026-06-03", "Admin", "", "開発", "", "実装", 2])  # in-file dup
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    files = {"file": ("wl.xlsx", _file(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r1 = auth_client.post("/api/worklog/import.xlsx", files=files)
    assert r1.status_code == 200, r1.text
    # One created, the identical second row skipped as duplicate.
    assert r1.json() == {"created": 1, "skipped": 0, "duplicates": 1}

    # Re-importing the same file adds nothing (both now duplicates of the stored log).
    files2 = {"file": ("wl.xlsx", _file(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = auth_client.post("/api/worklog/import.xlsx", files=files2)
    assert r2.json() == {"created": 0, "skipped": 0, "duplicates": 2}

    logs = auth_client.get("/api/worklog?from=2026-06-03&to=2026-06-03").json()
    assert len([l for l in logs if l["memo"] == "実装"]) == 1


def test_export_import_follows_configured_category_levels(auth_client):
    """分類の段数・名称は組織設定で変えられる: 見出しがその名前になり、3段目も往復する。"""
    r = auth_client.patch(
        "/api/org",
        json={"settings": {"worklog": {"category_levels": ["業務", "工程", "詳細"]}}},
    )
    assert r.status_code == 200, r.text

    ws = load_workbook(
        io.BytesIO(auth_client.get("/api/worklog/export.xlsx?from=2026-06-05&to=2026-06-05").content)
    ).active
    assert [c.value for c in ws[1]] == ["日付", "ユーザー", "タスクID", "業務", "工程", "詳細", "メモ", "時間"]

    wb = Workbook()
    out = wb.active
    out.append(["日付", "ユーザー", "タスクID", "業務", "工程", "詳細", "メモ", "時間"])
    out.append(["2026-06-05", "Admin", "", "開発", "設計", "画面", "作図", 4])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = auth_client.post(
        "/api/worklog/import.xlsx",
        files={"file": ("wl.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    assert r.json()["created"] == 1

    log = auth_client.get("/api/worklog?from=2026-06-05&to=2026-06-05").json()[0]
    assert (log["cat1"], log["cat2"], log["cat3"]) == ("開発", "設計", "画面")


def test_inspect_worklog_previews_without_writing(auth_client):
    """取り込み前に結果を確認できる: 追加/スキップ/重複の件数と理由。DBは変わらない。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["2026年6月の実績"])  # title line
    ws.append(["日付", "ユーザー", "タスクID", "大分類", "中分類", "メモ", "時間"])
    ws.append(["2026-06-10", "Admin", "", "開発", "", "実装", 3])
    ws.append(["2026-06-10", "Admin", "", "開発", "", "実装", 3])  # in-file duplicate
    ws.append(["2026-06-10", "居ない人", "", "", "", "", 2])  # skipped
    ws.append(["2026-06-10", "Admin", "", "", "", "会議", "abc"])  # hours not numeric
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        "/api/worklog/import.xlsx/inspect",
        files={"file": ("wl.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["suggested_header_row"] == 2
    assert (body["created"], body["duplicates"], body["skipped"]) == (1, 1, 2)
    assert any("居ない人" in i["reason"] for i in body["issues"])
    fields = {f["key"]: f for f in body["fields"]}
    assert fields["user"]["index"] == 1 and fields["hours"]["index"] == 6
    assert fields["cat1"]["label"] == "大分類"
    # Nothing was written by the dry run.
    assert auth_client.get("/api/worklog?from=2026-06-10&to=2026-06-10").json() == []


def test_import_worklog_honors_mapping(auth_client):
    """見出し名が違うファイルでも、列の対応を指定すれば取り込める。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["実施日", "担当者", "作業内容", "工数(h)"])
    ws.append(["2026-06-11", "Admin", "打合せ", 1.5])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = auth_client.post(
        "/api/worklog/import.xlsx",
        files={"file": ("wl.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mapping": json.dumps({"date": 0, "user": 1, "memo": 2, "hours": 3})},
    )
    assert r.status_code == 200, r.text
    assert r.json()["created"] == 1
    log = auth_client.get("/api/worklog?from=2026-06-11&to=2026-06-11").json()[0]
    assert log["memo"] == "打合せ" and float(log["hours"]) == 1.5


def test_import_worklog_requires_user_and_hours_columns(auth_client):
    buf = io.BytesIO()
    wb = Workbook()
    wb.active.append(["日付", "メモ"])
    wb.save(buf)
    buf.seek(0)
    r = auth_client.post(
        "/api/worklog/import.xlsx",
        files={"file": ("wl.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 400, r.text


def test_import_worklog_requires_admin(client, org_admin, db):
    """A non-admin member is rejected (403)."""
    from app import models
    from app.security import hash_password

    member = models.User(
        org_id=org_admin["org_id"],
        email="m@t.local",
        name="Member",
        role="member",
        password_hash=hash_password("pw123456"),
    )
    db.add(member)
    db.commit()
    client.post("/api/auth/login", json={"email": "m@t.local", "password": "pw123456"})

    wb = Workbook()
    ws = wb.active
    ws.append(["日付", "ユーザー", "時間"])
    ws.append(["2026-06-02", "Member", 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        "/api/worklog/import.xlsx",
        files={"file": ("wl.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 403, r.text
