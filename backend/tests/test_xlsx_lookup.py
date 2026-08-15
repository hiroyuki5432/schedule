"""XLOOKUP / VLOOKUP → 参照(LOOKUP)列 と、Excel の「テーブル」で書かれた数式。

要望: 新規シート作成でExcel取り込み時にXLOOKUP使ってたらうまくLOOKUPを自動生成して
ほしい／テーブルだと全然取り込めなさそう。

参照列は「別のシートを引く」ものなので、**参照先が既にこのアプリにあるとき** だけ
自動で作れる。無いときは値として取り込み、理由を画面に返す — というところまでを見る。
"""
from __future__ import annotations

import io
import json

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _save(wb) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _master_workbook() -> io.BytesIO:
    """先に取り込んでおくマスタ（品番・品名・単価）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "マスタ"
    ws.append(["品番", "品名", "単価"])
    for i in range(1, 4):
        ws.append([f"P{i:03d}", f"部品{i}", 100 * i])
    return _save(wb)


def _detail_workbook(*, as_table: bool) -> io.BytesIO:
    """明細シート。品名列が XLOOKUP でマスタを引いている。

    `as_table` で書き方が変わる:
      * True  … Excel の「テーブル」— 構造化参照（`XLOOKUP([@品番], マスタ[品番], …)`）
      * False … ふつうの範囲参照（`XLOOKUP(A2, マスタ!$A:$A, …)`）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["品番", "数量", "品名", "金額"])
    for i in range(1, 4):
        row = i + 1
        if as_table:
            ws.append([
                f"P{i:03d}", i,
                "=XLOOKUP([@品番],マスタ[品番],マスタ[品名])",
                "=[@数量]*XLOOKUP([@品番],マスタ[品番],マスタ[単価])",
            ])
        else:
            ws.append([
                f"P{i:03d}", i,
                f"=XLOOKUP(A{row},マスタ!$A:$A,マスタ!$B:$B)",
                f"=B{row}*2",
            ])

    ms = wb.create_sheet("マスタ")
    ms.append(["品番", "品名", "単価"])
    for i in range(1, 4):
        ms.append([f"P{i:03d}", f"部品{i}", 100 * i])

    if as_table:
        ws.add_table(Table(displayName="明細", ref="A1:D4"))
        ms.add_table(Table(displayName="マスタ", ref="A1:C4"))
    return _save(wb)


def _import_master(client) -> int:
    r = client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("m.xlsx", _master_workbook(), _MEDIA)},
        data={"name": "マスタ", "has_week_grid": "false", "sheet_name": "マスタ"},
    )
    assert r.status_code == 201, r.text
    return r.json()["sheet_id"]


def _inspect(client, buf, **extra) -> dict:
    r = client.post(
        "/api/sheets/import.xlsx/inspect",
        files={"file": ("in.xlsx", buf, _MEDIA)},
        data={"has_week_grid": "false", "sheet_name": "明細", **extra},
    )
    assert r.status_code == 200, r.text
    return {c["header"]: c for c in r.json()["columns"]}


# --------------------------------------------------------------------------- #
# 参照先がまだ無いとき
# --------------------------------------------------------------------------- #
def test_lookup_without_the_master_sheet_stays_a_value(auth_client):
    """マスタを取り込む前は参照列にできない — 理由を返し、値として取り込む。"""
    by_header = _inspect(auth_client, _detail_workbook(as_table=True))
    name = by_header["品名"]
    assert name["type"] != "lookup"
    assert name["formula"]["lookup"]["ready"] is False
    assert "マスタ" in name["formula"]["reason"]


# --------------------------------------------------------------------------- #
# 参照先があるとき（テーブルの書き方 / ふつうの範囲参照）
# --------------------------------------------------------------------------- #
def test_xlookup_over_a_table_becomes_a_lookup_column(auth_client):
    master_id = _import_master(auth_client)
    by_header = _inspect(auth_client, _detail_workbook(as_table=True))

    name = by_header["品名"]
    assert name["type"] == "lookup"          # 既定で参照列として提案される
    lk = name["formula"]["lookup"]
    assert lk["ready"] is True
    assert lk["sheet_id"] == master_id
    assert lk["local_column"] == "品番"
    assert lk["return_column"] == "品名"

    # 取り込んだマスタは1列目をID列にしているので、照合先は列ではなく行のID。
    assert lk["match_key_column_id"] == "__id__"

    # テーブルの中の掛け算（構造化参照）も数式列になる — ここが従来まったく読めなかった。
    money = by_header["金額"]
    assert money["type"] != "lookup"
    assert money["formula"]["reason"] is not None  # XLOOKUP を含む式は式にできない


def test_xlookup_over_plain_ranges_becomes_a_lookup_column(auth_client):
    master_id = _import_master(auth_client)
    by_header = _inspect(auth_client, _detail_workbook(as_table=False))

    lk = by_header["品名"]["formula"]["lookup"]
    assert lk["ready"] is True and lk["sheet_id"] == master_id
    # 列番地（$B:$B）でも、参照先の見出し行を読んで列名を突き止める。
    assert lk["return_column"] == "品名"
    assert by_header["金額"]["formula"]["expr"] == "[数量]*2"


def test_import_creates_the_lookup_column(auth_client):
    master_id = _import_master(auth_client)
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("in.xlsx", _detail_workbook(as_table=True), _MEDIA)},
        data={"name": "明細", "has_week_grid": "false", "sheet_name": "明細"},
    )
    assert r.status_code == 201, r.text
    detail = auth_client.get(f"/api/sheets/{r.json()['sheet_id']}").json()

    name = next(c for c in detail["columns"] if c["name"] == "品名")
    assert name["type"] == "lookup"
    cfg = name["config"]
    assert cfg["target_sheet_id"] == master_id
    assert cfg["match_key_column_id"] == "__id__"
    # キーは「このシートの品番列」。ID列（=品番）にしているので行のID。
    assert cfg["local_key_column_id"] == "__id__"

    # 参照列は計算列なので値を保存しない。
    assert all(str(name["id"]) not in row["data"] for row in detail["rows"])


def test_lookup_column_can_be_forced_back_to_a_value(auth_client):
    """ウィザードで「自由入力として取り込む」を選んだとき（lookup を送らない）。"""
    _import_master(auth_client)
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("in.xlsx", _detail_workbook(as_table=True), _MEDIA)},
        data={
            "name": "明細2",
            "has_week_grid": "false",
            "sheet_name": "明細",
            "columns": json.dumps(
                [
                    {"index": 1, "name": "数量", "type": "number"},
                    {"index": 2, "name": "品名", "type": "text"},
                ]
            ),
        },
    )
    assert r.status_code == 201, r.text
    detail = auth_client.get(f"/api/sheets/{r.json()['sheet_id']}").json()
    name = next(c for c in detail["columns"] if c["name"] == "品名")
    assert name["type"] == "text"


# --------------------------------------------------------------------------- #
# 名前がずれているとき — ウィザードで手で結びつける
# --------------------------------------------------------------------------- #
def _renamed_master(client) -> tuple[int, str]:
    """Excel の「マスタ」とは名前の違うシート＋列を用意する。"""
    r = client.post("/api/sheets", json={"name": "部品マスター", "has_week_grid": False})
    sid = r.json()["id"]
    c = client.post(
        f"/api/sheets/{sid}/columns", json={"name": "品目名称", "type": "text"}
    )
    return sid, str(c.json()["id"])


def test_a_name_mismatch_is_reported_not_guessed(auth_client):
    """似ているだけの名前を勝手に当てない — 理由を出して、選んでもらう。"""
    _renamed_master(auth_client)
    by_header = _inspect(auth_client, _detail_workbook(as_table=True))
    lk = by_header["品名"]["formula"]["lookup"]
    assert lk["ready"] is False
    # Excel 側が何と書いていたかは返す（画面で対応づけるときの手がかり）。
    assert lk["target_worksheet"] == "マスタ"
    assert lk["match_column"] == "品番" and lk["return_column"] == "品名"


def test_a_hand_picked_target_is_used(auth_client):
    """ウィザードで選んだ参照先が、そのまま参照列になる。"""
    sid, name_col = _renamed_master(auth_client)
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("in.xlsx", _detail_workbook(as_table=True), _MEDIA)},
        data={
            "name": "明細3",
            "has_week_grid": "false",
            "sheet_name": "明細",
            "columns": json.dumps(
                [
                    {"index": 1, "name": "数量", "type": "number"},
                    {
                        "index": 2,
                        "name": "品名",
                        "type": "lookup",
                        "lookup": {
                            "sheet_id": sid,
                            "local_index": 0,      # Excel の A列（＝ID列）で引く
                            "match_key_column_id": "__id__",
                            "return_column_id": name_col,
                        },
                    },
                ]
            ),
        },
    )
    assert r.status_code == 201, r.text
    detail = auth_client.get(f"/api/sheets/{r.json()['sheet_id']}").json()
    col = next(c for c in detail["columns"] if c["name"] == "品名")
    assert col["type"] == "lookup"
    assert col["config"] == {
        "target_sheet_id": sid,
        "local_key_column_id": "__id__",
        "match_key_column_id": "__id__",
        "return_column_id": name_col,
    }


def test_a_target_column_that_does_not_exist_is_refused(auth_client):
    """画面が言っているだけの指定は信用しない — 実在しない列なら参照列にしない。"""
    sid, _name_col = _renamed_master(auth_client)
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("in.xlsx", _detail_workbook(as_table=True), _MEDIA)},
        data={
            "name": "明細4",
            "has_week_grid": "false",
            "sheet_name": "明細",
            "columns": json.dumps(
                [
                    {
                        "index": 2,
                        "name": "品名",
                        "type": "lookup",
                        "lookup": {
                            "sheet_id": sid,
                            "local_index": 0,
                            "match_key_column_id": "__id__",
                            "return_column_id": "999999",   # そんな列は無い
                        },
                    }
                ]
            ),
        },
    )
    assert r.status_code == 201, r.text
    detail = auth_client.get(f"/api/sheets/{r.json()['sheet_id']}").json()
    col = next(c for c in detail["columns"] if c["name"] == "品名")
    # 参照列にはせず、値の列として取り込む（Excel にあった値まで消さない）。
    assert col["type"] != "lookup"


# --------------------------------------------------------------------------- #
# ファイルの中にだけある目印（_xlfn.）
# --------------------------------------------------------------------------- #
def test_xlfn_prefixed_xlookup_is_read_and_not_shown(auth_client):
    """Excel は XLOOKUP を `_xlfn.XLOOKUP` としてファイルに書く。

    画面には一度もそう出ないので、読むときは無視し、見せるときも外す。
    """
    master_id = _import_master(auth_client)
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["品番", "品名"])
    for i in range(1, 4):
        ws.append([f"P{i:03d}", f"=_xlfn.XLOOKUP(A{i + 1},マスタ!$A:$A,マスタ!$B:$B)"])
    ms = wb.create_sheet("マスタ")
    ms.append(["品番", "品名", "単価"])
    for i in range(1, 4):
        ms.append([f"P{i:03d}", f"部品{i}", 100 * i])

    by_header = _inspect(auth_client, _save(wb))
    info = by_header["品名"]
    assert info["type"] == "lookup"
    assert info["formula"]["lookup"]["sheet_id"] == master_id
    # 「_xlfn.」は利用者に見せない。
    assert info["formula"]["sample"] == "=XLOOKUP(A2,マスタ!$A:$A,マスタ!$B:$B)"


# --------------------------------------------------------------------------- #
# 自動では分解できない XLOOKUP / VLOOKUP
#
# 要望（不具合報告）: XLOOKUP も参照が選べるものと選べないものがある。なぜ？
#
# 原因は、画面が「参照先を選ぶ…」を出す条件を **分解できたか** にしていたこと。式の形が
# 想定外だと、ボタンごと消えるうえ理由も出ないので、利用者からは説明のつかない差にしか
# 見えない。いまは「XLOOKUP/VLOOKUP が書いてある列」なら必ず `has_lookup` を立て、
# **なぜ自動で結びつかなかったか** を `reason` で返す。ここではそれを全ケース見る。
# --------------------------------------------------------------------------- #
def _formula_workbook(formulas: list[str]) -> io.BytesIO:
    """明細シートの「品名」列（C列）に、行ごとの数式を並べたブック。

    `formulas` は2行目からの中身。`{row}` はその行の番号に置き換わる。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["品番", "数量", "品名"])
    for i, f in enumerate(formulas):
        ws.append([f"P{i + 1:03d}", i + 1, f.format(row=i + 2)])

    ms = wb.create_sheet("マスタ")
    ms.append(["品番", "品名", "単価"])
    for i in range(1, 4):
        ms.append([f"P{i:03d}", f"部品{i}", 100 * i])
    return _save(wb)


def _reason_for(client, formulas: list[str]) -> str:
    """「品名」列の formula ブロックを見て、必ずダイアログを開ける状態か確かめる。

    返すのは画面に出る理由（＝なぜ自動で結びつかなかったか）。
    """
    info = _inspect(client, _formula_workbook(formulas))["品名"]
    fx = info["formula"]
    # ここが今回の肝: 分解できなくても「参照先を選ぶ…」の入口は出す。
    assert fx["has_lookup"] is True, fx
    # 分解できていないので、自動で参照列にはしない（値として取り込む）。
    assert info["type"] != "lookup"
    assert fx["reason"], fx
    return fx["reason"]


_PLAIN = "=XLOOKUP(A{row},マスタ!$A:$A,マスタ!$B:$B)"


def test_lookup_inside_a_bigger_expression_is_offered_by_hand(auth_client):
    """1. XLOOKUP が式の一部（掛け算・文字列連結）。"""
    for f in (
        "=B{row}*XLOOKUP(A{row},マスタ!$A:$A,マスタ!$C:$C)",
        '=XLOOKUP(A{row},マスタ!$A:$A,マスタ!$B:$B)&"個"',
    ):
        reason = _reason_for(auth_client, [f, f, f])
        assert "式の一部" in reason, reason
        # 数式列として落ちたときの「対応していない関数です：XLOOKUP」は、XLOOKUP を
        # 書いた本人にはまるで的外れ。もう出さない。
        assert "対応していない関数" not in reason


def test_a_column_with_only_some_lookup_rows_says_so(auth_client):
    """2. 列の一部の行だけ XLOOKUP。"""
    reason = _reason_for(auth_client, [_PLAIN, "=B{row}*2", _PLAIN])
    assert "一部の行" in reason and "3行中2行" in reason, reason
    assert "対応していない関数" not in reason


def test_a_key_that_is_not_a_plain_cell_says_so(auth_client):
    """3. 探す値が素のセルでない（連結・関数を通している）。"""
    for f in (
        "=XLOOKUP(A{row}&B{row},マスタ!$A:$A,マスタ!$B:$B)",
        "=XLOOKUP(TRIM(A{row}),マスタ!$A:$A,マスタ!$B:$B)",
    ):
        reason = _reason_for(auth_client, [f, f, f])
        assert "探す値が同じ行の列ではありません" in reason, reason


def test_non_exact_match_modes_say_so(auth_client):
    """4. 照合モードが 0 以外 / 末尾から検索 / VLOOKUP の第4引数が TRUE・省略。"""
    cases = {
        '=XLOOKUP(A{row},マスタ!$A:$A,マスタ!$B:$B,"",1)': "完全一致",
        '=XLOOKUP(A{row},マスタ!$A:$A,マスタ!$B:$B,"",0,-1)': "末尾から探す",
        "=VLOOKUP(A{row},マスタ!$A:$C,2,TRUE)": "完全一致",
        "=VLOOKUP(A{row},マスタ!$A:$C,2)": "完全一致",
    }
    for f, expected in cases.items():
        reason = _reason_for(auth_client, [f, f, f])
        assert expected in reason, (f, reason)


def test_an_unreadable_range_says_so(auth_client):
    """5. 照合範囲が1列でない / 照合と取得が別ワークシート。"""
    reason = _reason_for(auth_client, ["=XLOOKUP(A{row},マスタ!$A:$B,マスタ!$C:$C)"] * 3)
    assert "1列ではありません" in reason, reason

    reason = _reason_for(auth_client, ["=XLOOKUP(A{row},マスタ!$A:$A,マスタ2!$B:$B)"] * 3)
    assert "別のワークシート" in reason, reason


def test_rows_pointing_at_different_targets_say_so(auth_client):
    """6. 行によって参照先が違う（品名を引く行と単価を引く行）。"""
    reason = _reason_for(
        auth_client,
        [_PLAIN, "=XLOOKUP(A{row},マスタ!$A:$A,マスタ!$C:$C)", _PLAIN],
    )
    assert "行によって参照先が違う" in reason, reason


def test_a_resolvable_lookup_also_reports_has_lookup(auth_client):
    """分解できた列でも `has_lookup` は立つ（ボタンの出し分けを1つの条件で済ませる）。"""
    _import_master(auth_client)
    fx = _inspect(auth_client, _detail_workbook(as_table=True))["品名"]["formula"]
    assert fx["has_lookup"] is True and fx["lookup"]["ready"] is True


def test_a_plain_formula_column_has_no_lookup_entry_point(auth_client):
    """LOOKUP と関係ない数式の列に、参照先の入口を出したりはしない。"""
    fx = _inspect(auth_client, _formula_workbook(["=B{row}*2"] * 3))["品名"]["formula"]
    assert fx["has_lookup"] is False
    assert fx["expr"] == "[数量]*2"


# --------------------------------------------------------------------------- #
# テーブルで書かれた、ふつうの数式
# --------------------------------------------------------------------------- #
def _table_arithmetic_workbook() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "見積"
    ws.append(["ID", "単価", "数量", "金額"])
    for i in range(1, 4):
        ws.append([f"K-{i}", 100 * i, i, "=[@単価]*[@数量]"])
    ws.add_table(Table(displayName="見積T", ref="A1:D4"))
    return _save(wb)


def test_structured_reference_arithmetic_becomes_a_formula_column(auth_client):
    """`=[@単価]*[@数量]` — テーブルの中の数式。従来はここで丸ごと落ちていた。"""
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("t.xlsx", _table_arithmetic_workbook(), _MEDIA)},
        data={"name": "見積T", "has_week_grid": "false"},
    )
    assert r.status_code == 201, r.text
    detail = auth_client.get(f"/api/sheets/{r.json()['sheet_id']}").json()
    money = next(c for c in detail["columns"] if c["name"] == "金額")
    assert money["type"] == "formula"
    assert money["config"]["expr"] == "[単価]*[数量]"


# --------------------------------------------------------------------------- #
# 「ready と言いながら違う参照になる」たぐいの回帰（検証で見つかったもの）
# --------------------------------------------------------------------------- #
def _lookup_of(client, formula: str, *, master_rows=None):
    """明細の1列が `formula` だけのブックを inspect して、その列の情報を返す。

    `master_rows` は `(見出しの並び, 行の並び)`。見出しと中身がちぐはぐだと
    「その列が行IDになっているか」の判定を正しく試せないので、まとめて渡す。
    """
    headers, rows = master_rows or (
        ["品番", "品名", "単価"],
        [[f"P{i:03d}", f"部品{i}", 100 * i] for i in range(1, 4)],
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["品番", "品名"])
    for i in range(1, 4):
        ws.append([f"P{i:03d}", formula.replace("{row}", str(i + 1))])
    ms = wb.create_sheet("マスタ")
    ms.append(list(headers))
    for row in rows:
        ms.append(list(row))
    return _inspect(client, _save(wb))["品名"]


def test_a_misaligned_range_is_refused(auth_client):
    """`XLOOKUP(A2, マスタ!$A$2:$A$10, マスタ!$B$3:$B$11)` は1行ずれた値を返す式。

    行番号を読み捨てていたときは「揃った参照」として通してしまっていた。
    """
    _import_master(auth_client)
    info = _lookup_of(auth_client, "=XLOOKUP(A{row},マスタ!$A$2:$A$10,マスタ!$B$3:$B$11)")
    assert info["type"] != "lookup"
    assert "ずれ" in info["formula"]["reason"]
    # 手で選ぶ道は残る。
    assert info["formula"]["has_lookup"] is True


def test_an_iferror_falling_back_to_another_lookup_is_refused(auth_client):
    """逃げ先が別の LOOKUP なら、2つめを黙って捨ててはいけない。"""
    _import_master(auth_client)
    info = _lookup_of(
        auth_client,
        "=IFERROR(XLOOKUP(A{row},マスタ!$A:$A,マスタ!$B:$B),"
        "XLOOKUP(A{row},マスタ!$A:$A,マスタ!$C:$C))",
    )
    assert info["type"] != "lookup"
    assert info["formula"]["reason"]
    assert info["formula"]["has_lookup"] is True


def test_the_id_column_guess_is_checked_against_the_data(auth_client):
    """「マスタの先頭列が ID 列」と決めつけない — 値が重なることを確かめる。

    マスタを **2列目** を ID 列にして取り込むと、先頭列（社内メモ）は取り込まれない。
    そこを XLOOKUP が照合していても、行ID と重ならないので参照列にはできない。
    以前は見出し名だけで `__id__` に落としており、「社内メモで照合」と書かれた式が
    黙って「品番の行IDで照合」する参照列になっていた。
    """
    # 社内メモ / 品番 / 品名 の3列。品番（index=1）を ID 列にして取り込む。
    headers = ["社内メモ", "品番", "品名"]
    rows = [[f"メモ{i}", f"P{i:03d}", f"部品{i}"] for i in range(1, 4)]
    wb = Workbook()
    ws = wb.active
    ws.title = "マスタ"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("m.xlsx", _save(wb), _MEDIA)},
        data={
            "name": "マスタ", "has_week_grid": "false", "sheet_name": "マスタ",
            "id_column": "1",
            "columns": json.dumps([{"index": 2, "name": "品名", "type": "text"}]),
        },
    )
    assert r.status_code == 201, r.text

    # 明細は A列（社内メモ）で照合する式。行ID は品番の値なので、社内メモとは重ならない。
    info = _lookup_of(
        auth_client,
        "=XLOOKUP(A{row},マスタ!$A:$A,マスタ!$C:$C)",
        master_rows=(headers, rows),
    )
    lk = info["formula"]["lookup"]
    assert lk["match_column"] == "社内メモ"
    assert lk["ready"] is False          # 勝手に __id__ に落とさない
    assert lk["match_key_column_id"] == ""
    assert info["type"] != "lookup"


def test_the_real_id_column_is_still_found(auth_client):
    """値が重なっていれば、これまでどおり行ID（__id__）に結びつく。

    上の検査だけだと「常に諦める」実装でも通ってしまうので、逆側も見る。
    """
    _import_master(auth_client)   # 品番を ID 列にして取り込む（既定）
    info = _lookup_of(auth_client, "=XLOOKUP(A{row},マスタ!$A:$A,マスタ!$B:$B)")
    lk = info["formula"]["lookup"]
    assert lk["ready"] is True
    assert lk["match_key_column_id"] == "__id__"


def test_a_formula_referring_to_the_id_column_uses_ID(auth_client):
    """ID列を指す式は `[ID]` に翻訳する。

    取り込むと ID列 は「シートの列」ではなく行のID（key_value）になるので、見出しの
    名前で参照する式を作ると存在しない列を指し、列まるごと
    `#「品番」という列がありません` になっていた。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["品番", "連絡先"])
    for i in range(1, 4):
        ws.append([f"P{i:03d}", f'=A{i + 1}&"@example.com"'])
    info = _inspect(auth_client, _save(wb))["連絡先"]
    assert info["type"] == "formula"
    assert info["formula"]["expr"] == '[ID]&"@example.com"'

    r = auth_client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("in.xlsx", _save(wb), _MEDIA)},
        data={"name": "連絡先テスト", "has_week_grid": "false", "sheet_name": "明細"},
    )
    assert r.status_code == 201, r.text
    detail = auth_client.get(f"/api/sheets/{r.json()['sheet_id']}").json()
    col = next(c for c in detail["columns"] if c["name"] == "連絡先")
    assert col["config"]["expr"] == '[ID]&"@example.com"'


# --------------------------------------------------------------------------- #
# 「式は作られるが、参照先の列が無い」たぐいの回帰
#
# どれも列まるごとが `#「◯◯」という列がありません` になっていた。式を作ったときの
# 名前と、実際に作られる列名を突き合わせる場所が無かったのが原因。
# --------------------------------------------------------------------------- #
def _priced_book():
    """品番 / 数量 / 単価 / 金額(=数量*単価)。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["品番", "数量", "単価", "金額"])
    for i in range(1, 4):
        r = i + 1
        ws.append([f"P{i:03d}", i, 100 * i, f"=B{r}*C{r}"])
    return wb


def _import_cols(client, name, columns):
    r = client.post(
        "/api/sheets/import.xlsx",
        files={"file": ("a.xlsx", _save(_priced_book()), _MEDIA)},
        data={"name": name, "has_week_grid": "false", "columns": json.dumps(columns)},
    )
    assert r.status_code == 201, r.text
    d = client.get(f"/api/sheets/{r.json()['sheet_id']}").json()
    return {c["name"]: c for c in d["columns"]}


def test_a_renamed_column_keeps_the_formula_valid(auth_client):
    """ウィザードで「単価」を「価格」に打ち直したら、式もそちらを指す。"""
    ins = auth_client.post(
        "/api/sheets/import.xlsx/inspect",
        files={"file": ("a.xlsx", _save(_priced_book()), _MEDIA)},
        data={"has_week_grid": "false", "columns": json.dumps([
            {"index": 1, "name": "数量", "type": "number"},
            {"index": 2, "name": "価格", "type": "number"},
            {"index": 3, "name": "金額", "type": "formula"},
        ])},
    ).json()
    money = next(c for c in ins["columns"] if c["header"] == "金額")
    assert money["formula"]["expr"] == "[数量]*[価格]"

    cols = _import_cols(auth_client, "改名", [
        {"index": 1, "name": "数量", "type": "number"},
        {"index": 2, "name": "価格", "type": "number"},
        {"index": 3, "name": "金額", "type": "formula", "expr": "[数量]*[価格]"},
    ])
    assert cols["金額"]["config"]["expr"] == "[数量]*[価格]"


def test_a_formula_over_an_unchecked_column_is_refused(auth_client):
    """単価のチェックを外したら、それを参照する式は作らない（値として取り込む）。"""
    ins = auth_client.post(
        "/api/sheets/import.xlsx/inspect",
        files={"file": ("a.xlsx", _save(_priced_book()), _MEDIA)},
        data={"has_week_grid": "false", "columns": json.dumps([
            {"index": 1, "name": "数量", "type": "number"},
            {"index": 3, "name": "金額", "type": "formula"},
        ])},
    ).json()
    money = next(c for c in ins["columns"] if c["header"] == "金額")
    assert money["formula"]["expr"] is None
    assert money["formula"]["reason"]        # 理由が画面に出る

    # 古い式を送りつけられても、参照先が無ければ数式列にしない。
    cols = _import_cols(auth_client, "外し", [
        {"index": 1, "name": "数量", "type": "number"},
        {"index": 3, "name": "金額", "type": "formula", "expr": "[数量]*[単価]"},
    ])
    assert cols["金額"]["type"] != "formula"


def test_a_column_named_ID_blocks_the_id_reference(auth_client):
    """`ID` という列があると `[ID]` は行のキーではなくその列を指す。翻訳しない。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["品番", "ID", "表示"])
    for i in range(1, 4):
        ws.append([f"P{i:03d}", f"X{i}", f'=A{i + 1}&"-"&B{i + 1}'])
    ins = _inspect(auth_client, _save(wb))
    assert ins["表示"]["formula"]["expr"] is None
    assert "ID" in ins["表示"]["formula"]["reason"]


def test_a_partly_manual_column_is_not_a_formula_by_default(auth_client):
    """5行中2行だけ数式。既定で数式列にすると手入力の3行が消える。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["品番", "数量", "単価", "金額"])
    for i in range(1, 6):
        r = i + 1
        ws.append([f"P{i:03d}", i, 100 * i,
                   f"=B{r}*C{r}" if i <= 2 else 99999])
    info = _inspect(auth_client, _save(wb))["金額"]
    assert info["type"] != "formula"                    # 既定にしない
    assert info["formula"]["covers_all_rows"] is False
    assert info["formula"]["replaced_values"] == 3      # 消える行数を画面に出せる


def test_xlookup_with_a_formula_fallback_is_refused(auth_client):
    """第4引数（見つからないときの値）に式が書かれていたら参照列にしない。"""
    from app import xlsx_formula as F
    import pytest as _pytest

    with _pytest.raises(F.Untranslatable):
        F.extract_lookup(
            '=XLOOKUP(A2,マスタ!$A:$A,マスタ!$B:$B,XLOOKUP(A2,マスタ!$A:$A,マスタ!$C:$C))',
            2, {0: "品番"}, worksheet="明細", tables={},
        )
    # 定数の逃げ先はこれまでどおり通る。
    spec = F.extract_lookup(
        '=XLOOKUP(A2,マスタ!$A:$A,マスタ!$B:$B,"未登録")',
        2, {0: "品番"}, worksheet="明細", tables={},
    )
    assert spec.target_worksheet == "マスタ"


def test_two_sheets_with_the_same_name_are_not_guessed(auth_client):
    """同名シートが複数あるなら、どれかを黙って選ばない。

    画面には名前しか出ないので、どちらに繋がったのか区別がつかない。並び順の保証も
    無いので、同じファイルを流し直すと別のシートを指す参照列ができうる。
    """
    for _ in range(2):
        r = auth_client.post("/api/sheets", json={"name": "マスタ", "has_week_grid": False})
        assert r.status_code in (200, 201), r.text
        auth_client.post(
            f"/api/sheets/{r.json()['id']}/columns", json={"name": "品名", "type": "text"}
        )
    lk = _inspect(auth_client, _detail_workbook(as_table=True))["品名"]["formula"]["lookup"]
    assert lk["ready"] is False
    assert "複数あります" in (lk["reason"] or "")
