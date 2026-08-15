"""Shared .xlsx reading helpers for the 取り込みウィザード.

Every import in the app follows the same shape: open the workbook, let the user
pick a worksheet and the 見出し行, show a raw preview, then map columns before
anything is written. The generic parts live here so the sheet importer
(``routers.excel``) and the 日報 importer (``routers.worklog``) behave the same.
"""
from __future__ import annotations

import io
import json
from datetime import date, datetime

from fastapi import HTTPException, UploadFile, status
from app.date_values import parse_date_value

#: Raw rows returned for the 見出し行 picker.
PREVIEW_ROWS = 30
#: Raw rows returned from the BOTTOM, for the 「これ以降を取り込まない」 picker.
TAIL_PREVIEW_ROWS = 15
#: Hard cap on one tail window, so walking up a 5000-row sheet stays cheap.
TAIL_PREVIEW_MAX = 300
#: Rows kept visible ABOVE the current cut, so a typed 最終行 can be eyeballed.
TAIL_CONTEXT_ROWS = 8
#: How far down we look when guessing the header row.
HEADER_SCAN_ROWS = 10
#: Sample values shown per column.
SAMPLE_LIMIT = 4
#: Rows sampled when inferring a column's type / listing its values.
SCAN_ROWS = 200


def workbook_from_bytes(data: bytes, *, data_only: bool = True):
    """開いたワークブック。

    `data_only=True` は数式セルを **Excel が最後に保存した計算結果** として返す（通常の
    取り込みはこちら）。`data_only=False` にすると数式そのもの（`'=C2*D2'`）が返るので、
    数式列として取り込めるかの判定に使う（:mod:`app.xlsx_formula`）。
    """
    from openpyxl import load_workbook

    try:
        return load_workbook(io.BytesIO(data), data_only=data_only, read_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excelファイルを読み込めませんでした（.xlsx 形式をご確認ください）",
        )


def upload_bytes(file: UploadFile) -> bytes:
    """The uploaded file's bytes — read ONCE.

    An UploadFile's stream is consumed by the first read, so anything that needs
    the same workbook for several worksheets (一括取り込み) must hold these bytes
    and build its workbook from them instead of re-reading the upload.
    """
    return file.file.read()


def open_workbook(file: UploadFile):
    return workbook_from_bytes(upload_bytes(file))


def pick_worksheet(wb, sheet_name: str):
    """The named worksheet, or the active one when no name is given."""
    nm = (sheet_name or "").strip()
    if not nm:
        return wb.active
    if nm not in wb.sheetnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"ワークシート「{nm}」が見つかりません",
        )
    return wb[nm]


def grid_of(ws) -> list[tuple]:
    return list(ws.iter_rows(values_only=True))


def is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def auto_header_row(grid: list[tuple]) -> int:
    """1-based guess at the header row: the topmost of the first rows with the most
    filled cells — files often open with a title line or a blank row or two."""
    best_i, best_n = 1, -1
    for i, row in enumerate(grid[:HEADER_SCAN_ROWS], start=1):
        n = sum(0 if is_blank(v) else 1 for v in row)
        if n > best_n:
            best_i, best_n = i, n
    return best_i


def split_grid(grid: list[tuple], header_row: int, last_row: int = 0) -> tuple[tuple, list[tuple]]:
    """(header tuple, data rows below it) — fully empty data lines dropped.

    `last_row` is the 1-based worksheet row to stop AT (inclusive); 0 = 最後まで.
    Sheets very often end with something that is not part of the table — a 合計
    row, notes, a second little table — and there is no reliable way to tell those
    from data, so the user draws the line and it is applied here, on the PHYSICAL
    rows, before blanks are dropped. That way the number they clicked in the
    preview is the number that takes effect.
    """
    header = grid[header_row - 1] if 0 < header_row <= len(grid) else ()
    end = last_row if 0 < last_row <= len(grid) else len(grid)
    body = [r for r in grid[header_row:end] if r is not None and any(not is_blank(v) for v in r)]
    return header, body


def normalize_text(s: str) -> str:
    """Clean up a text cell's line breaks.

    A cell filled in with Alt+Enter comes back with real newlines, which we keep —
    but the line ENDINGS vary by who wrote the file, and some writers leave the
    literal ``_x000D_`` escape for a carriage return in the shared-string table.
    Left alone those show up as visible junk and make two identical-looking values
    compare unequal, so everything is folded to plain ``\\n`` on the way in.
    """
    return s.replace("_x000D_", "").replace("\r\n", "\n").replace("\r", "\n")


def cell_text(v) -> str:
    """Display text for a preview cell."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat() if (v.hour, v.minute, v.second) == (0, 0, 0) else v.isoformat(sep=" ")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return normalize_text(v).strip() if isinstance(v, str) else str(v)


def column_values(data_rows: list[tuple], idx: int, limit: int = SCAN_ROWS) -> list:
    """The first `limit` values of one column (out-of-range → None)."""
    if idx < 0:
        return []
    return [r[idx] if idx < len(r) else None for r in data_rows[:limit]]


def cell_at(row: tuple, idx: int):
    return row[idx] if 0 <= idx < len(row) else None


def looks_numeric(v) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).strip())
        return True
    except (TypeError, ValueError):
        return False


def coerce_date(raw) -> date | None:
    """Excel cell → date object, or None when it isn't one.

    書き方の違い（`2025/10/18`・`2025年10月18日`・時刻つき）の吸収は
    :mod:`app.date_values` に集約している。
    """
    return parse_date_value(raw)


def preview_of(grid: list[tuple], limit: int = PREVIEW_ROWS) -> list[dict]:
    """First rows with their 1-based row numbers, for the 見出し行 picker."""
    return [
        {"row": i, "cells": [cell_text(v) for v in row]}
        for i, row in enumerate(grid[:limit], start=1)
    ]


def tail_preview_of(
    grid: list[tuple], header_row: int, tail_from: int = 0, last_row: int = 0
) -> list[dict]:
    """A window of rows ending at the bottom, for the 最終行 picker.

    The head preview stops at row 30, but the rows that need cutting off live at
    the BOTTOM — so without this the user cannot see, let alone click, the row
    they want to stop at. Never reaches above the 見出し行.

    The window is not fixed to the last few rows: the cut is often much further
    up (a second little table, a section that is out of scope), so

    - `tail_from` (1-based) opens the window at an explicit row — that is the
      「さらに上を表示」 button walking upward;
    - a `last_row` already set drags the window up on its own, so a row number
      TYPED into the 最終行 box can always be checked against the real rows
      instead of leaving the user cutting blind.

    Capped at TAIL_PREVIEW_MAX rows per call so a huge sheet stays cheap; the
    button just gets pressed again.
    """
    floor = header_row + 1  # never above the 見出し行
    start = tail_from if tail_from > 0 else len(grid) - TAIL_PREVIEW_ROWS + 1
    if last_row > 0:
        start = min(start, last_row - TAIL_CONTEXT_ROWS)
    start = max(floor, min(start, len(grid)))
    window = grid[start - 1 : start - 1 + TAIL_PREVIEW_MAX]
    return [
        {"row": i, "cells": [cell_text(v) for v in row]}
        for i, row in enumerate(window, start=start)
    ]


def data_row_total(grid: list[tuple], header_row: int) -> int:
    """How many data rows the worksheet has with NO cut — so the UI can say how
    many rows the 最終行 setting is actually excluding."""
    return len(split_grid(grid, header_row)[1])


def worksheets_of(wb) -> list[dict]:
    return [
        {"name": s.title, "rows": s.max_row or 0, "columns": s.max_column or 0}
        for s in wb.worksheets
    ]


def parse_json_field(raw: str, what: str):
    """A JSON form field (the wizard's mapping/selection), or None when omitted."""
    if not (raw or "").strip():
        return None
    try:
        return json.loads(raw)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=f"{what}の指定を読み取れませんでした"
        )


def slice_worksheet(wb, sheet_name: str, header_row: int, last_row: int = 0):
    """One worksheet of an already-open workbook, sliced the way every wizard step
    does: (worksheet, grid, resolved 1-based header row, header tuple, data rows).
    `last_row` (1-based, inclusive; 0 = 最後まで) drops everything below it.
    Raises 400 on an empty worksheet."""
    ws = pick_worksheet(wb, sheet_name)
    grid = grid_of(ws)
    if not grid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="空のファイルです")
    hr = min(header_row if header_row > 0 else auto_header_row(grid), len(grid))
    header, data_rows = split_grid(grid, hr, last_row)
    return ws, grid, hr, header, data_rows


def read_source(file: UploadFile, sheet_name: str, header_row: int, last_row: int = 0):
    """Open + slice a workbook the way every wizard step does.

    Returns (workbook, worksheet, grid, resolved 1-based header row, header tuple,
    data rows). Raises 400 on an empty file.
    """
    wb = open_workbook(file)
    ws, grid, hr, header, data_rows = slice_worksheet(wb, sheet_name, header_row, last_row)
    return wb, ws, grid, hr, header, data_rows


def read_source_with_formulas(
    file: UploadFile, sheet_name: str, header_row: int, last_row: int = 0
):
    """`read_source` に加えて、同じ範囲の **数式** と、ブックの **テーブル定義** も返す。

    値と数式は openpyxl の別々の読み込みでしか取れない（`data_only` は開くときに決まる）
    ので、アップロードのバイト列を1回だけ読んで2回開く。返り値の後ろ2つは

    * ``[(ワークシート行番号, 生セル), …]`` の行リスト（値側の `data_rows` と同じ行だけが
      同じ順で並ぶ）
    * ``{テーブル名: TableDef}``（`[@数量]` のような書き方を読むために要る。
      :mod:`app.xlsx_tables` 参照）
    """
    from app.xlsx_tables import read_tables

    data = upload_bytes(file)
    wb = workbook_from_bytes(data)
    ws, grid, hr, header, data_rows = slice_worksheet(wb, sheet_name, header_row, last_row)

    fwb = workbook_from_bytes(data, data_only=False)
    try:
        fws = pick_worksheet(fwb, ws.title)
        fgrid = grid_of(fws)
    except HTTPException:
        fgrid = []

    # 値側は「空行を落とした後」の並び。数式側も同じ行だけを、物理行番号つきで拾う。
    end = last_row if 0 < last_row <= len(grid) else len(grid)
    formula_rows: list[tuple[int, tuple]] = []
    for i in range(hr, end):  # i は 0 始まり = 物理行番号 - 1
        row = grid[i]
        if row is None or not any(not is_blank(v) for v in row):
            continue
        formula_rows.append((i + 1, fgrid[i] if i < len(fgrid) else ()))
    return wb, ws, grid, hr, header, data_rows, formula_rows, read_tables(data)


def worksheet_header_names(wb, sheet_name: str, header_row: int = 0) -> dict[int, str]:
    """**別の** ワークシートの見出し（列位置 → 見出し名）。

    XLOOKUP が `マスタ!$C:$C` のように列番地で書かれているとき、C列が何という列なのかは
    参照先の見出し行を読まないと分からない。見出しの手前だけ読めば足りるので、上から
    数行だけ見る。`header_row`（1始まり）が分からなければこれまでと同じ推測を使う。
    """
    try:
        ws = pick_worksheet(wb, sheet_name)
    except HTTPException:
        return {}
    rows = list(ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True))
    if not rows:
        return {}
    hr = header_row if 0 < header_row <= len(rows) else auto_header_row(rows)
    return {i: cell_text(v) for i, v in enumerate(rows[hr - 1]) if cell_text(v)}


def worksheet_column_values(
    wb, sheet_name: str, header_row: int, index: int, limit: int = SCAN_ROWS
) -> set[str]:
    """**別の** ワークシートの、1列ぶんの値（見出し行より下）。

    「この列が、取り込み先シートの行ID になっているか」を **値で確かめる** ために使う。
    見出しの名前だけでは、どの列を ID 列にして取り込んだのかは分からない。
    """
    try:
        ws = pick_worksheet(wb, sheet_name)
    except HTTPException:
        return set()
    if index < 0:
        return set()
    hr = header_row if header_row > 0 else 1
    out: set[str] = set()
    for row in ws.iter_rows(
        min_row=hr + 1, max_row=hr + limit, min_col=index + 1, max_col=index + 1,
        values_only=True,
    ):
        text = cell_text(row[0]) if row else ""
        if text:
            out.add(text)
    return out
