"""Excel (.xlsx) export & import for a sheet.

Layout (one worksheet), schedule sheets:
    ID | <attribute columns…(incl 開始日/完了日)> | 進捗(%) | 先行タスク(ID)
       | <◇予定/◇実績 per template milestone…> | <week_start ISO dates…>

- 開始日/完了日 are now real date columns, so they round-trip as ordinary
  attribute columns. Phases are NOT exported — their boundary is derived from the
  ◇ dates + 開始日 on import. Only the sheet's TEMPLATE milestones (既定マイルストン)
  get 予定/実績 columns.
- 進捗 / 先行タスク round-trip too (先行タスク by ID/key_value, resolved after all
  rows are imported, so a clear → re-import fully restores the schedule).
- Export renders human-readable values (member → name). 計算列（参照(LOOKUP)／数式）は
  自動計算なので、書き出しは空欄・取り込みは対象外。
- Import upserts by ID (key_value). Weekly cells: past weeks write 実績,
  current/future write 予定 (the same display rule as the grid / CSV export).
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from uuid import uuid4

from fastapi import APIRouter, Depends, Form, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app import xlsx_formula, xlsx_import as xlsx
from app.db import get_db
from app.deps import get_sheet_for_user
from app.models import Column, EffortEntry, ImportPreset, Row, RowMilestone, Sheet, User
from app.schedule_service import ensure_schedule_columns, sched_columns
from app.date_values import is_date_placeholder, normalize_date_text, parse_date_value
from app.schemas import COMPUTED_COLUMN_TYPES
from app.security import current_user
from app.weeks import current_week_start, week_start_of
from app.worklog_service import org_week_start_weekday

router = APIRouter(prefix="/api/sheets", tags=["excel"])

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PROGRESS_HEADER = "進捗(%)"
DEPS_HEADER = "先行タスク(ID)"

#: 参照(LOOKUP)の設定で「列ではなく行のID（key_value）」を指す合図。frontend の
#: `lib/lookup.ts` の ID_KEY と同じ値でなければならない。
LOOKUP_ID_KEY = "__id__"

#: 取り込みで「行をどう照合するか」。
#:   'none'    … 照合しない。Excel の1行＝アプリの1行（既定）。同じIDが並んでいても
#:                別々の行として入る。ID列を指定していなければ内部IDを自動採番する。
#:   'id'      … ID列で既存の行を探して更新する（従来の動き）。同じIDは1行になる。
#:   'replace' … 取り込む前にシートの行を全部消してから入れる（入れ替え）。
MATCH_MODES = ("none", "id", "replace")


def resolve_match_mode(mode: str, id_column: int) -> str:
    """指定された照合モード。未指定（''）は従来の意味に解決する。

    この設定より前に保存されたプリセットや、モードを送ってこない API 利用は
    「ID列があるなら ID で照合」という従来の動きのままにする — 黙って挙動が変わって
    再取り込みが行を増殖させる、という壊れ方をしないため。画面（ウィザード）は
    常に明示的に送るので、そちらの既定は 'none'。
    """
    m = (mode or "").strip()
    return m if m in MATCH_MODES else ("id" if id_column >= 0 else "none")


# --------------------------------------------------------------------------- #
# Template milestone columns
# --------------------------------------------------------------------------- #
def _template_items(sheet: Sheet) -> list[dict]:
    """The sheet's ordered phase/milestone template (既定マイルストン)."""
    return list((sheet.settings or {}).get("default_milestones") or [])


def _template_milestone_cols(sheet: Sheet) -> list[tuple[int, str, str, str]]:
    """For each ◇ in the template: (template_index, name, 予定_header, 実績_header).
    Duplicate names get a ``#N`` suffix so headers stay unique."""
    out: list[tuple[int, str, str, str]] = []
    seen: dict[str, int] = {}
    for i, it in enumerate(_template_items(sheet)):
        if (it.get("kind") or "phase") != "milestone":
            continue
        name = (it.get("name") or "").strip() or f"◇{i + 1}"
        seen[name] = seen.get(name, 0) + 1
        label = name if seen[name] == 1 else f"{name} #{seen[name]}"
        out.append((i, name, f"{label}（予定）", f"{label}（実績）"))
    return out


#: Excel cell → date object, or None. 書き方の違いは app.date_values に集約。
_coerce_date_obj = parse_date_value


def _reconstruct_milestones(
    items: list[dict],
    predicted: dict[int, date],
    actual: dict[int, date],
    start_iso: str | None,
) -> list[dict]:
    """Rebuild a row's milestones from the template + the imported ◇ dates.
    Phases are recreated with boundaries derived from the preceding ◇ (first → 開始日)."""
    parsed: list[dict] = []
    for i, it in enumerate(items):
        kind = it.get("kind") or "phase"
        name = (it.get("name") or "").strip()
        if kind == "milestone":
            pred = predicted.get(i)
            if pred is None:
                continue  # no planned date for this ◇ → omit it
            parsed.append(
                {
                    "kind": "milestone",
                    "name": name,
                    "boundary_date": pred,
                    "actual_date": actual.get(i),
                    "done": actual.get(i) is not None,
                    "order": i,
                }
            )
        else:
            parsed.append(
                {"kind": "phase", "name": name, "boundary_date": None, "done": False, "actual_date": None, "order": i}
            )

    # Derive phase boundaries: forward-fill from the preceding ◇ (seeded with 開始日),
    # then back-fill any leading phases from the next dated entry.
    try:
        last = date.fromisoformat(start_iso) if start_iso else None
    except ValueError:
        last = None
    for p in parsed:
        if p["kind"] == "milestone":
            last = p["boundary_date"]
        elif p["boundary_date"] is None:
            p["boundary_date"] = last
    nxt = None
    for p in reversed(parsed):
        if p["boundary_date"] is None:
            p["boundary_date"] = nxt
        else:
            nxt = p["boundary_date"]
    return [p for p in parsed if p["boundary_date"] is not None]


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _members_by_id(db: Session, org_id: int) -> dict[str, str]:
    rows = db.execute(select(User).where(User.org_id == org_id)).scalars()
    return {str(u.id): u.name for u in rows}


def _members_by_name(db: Session, org_id: int) -> dict[str, int]:
    rows = db.execute(select(User).where(User.org_id == org_id)).scalars()
    # First match wins on duplicate names.
    out: dict[str, int] = {}
    for u in rows:
        out.setdefault(u.name, u.id)
    return out


def _fmt_num(value) -> float | str:
    if value is None:
        return ""
    f = float(value)
    return int(f) if f == int(f) else f


def _span_weeks(
    rows: list[Row], columns: list[Column], week_weekday: int
) -> list[date]:
    """Week columns to OUTPUT so 工数 can be entered even before any effort exists.

    Spans the rows' 開始日〜完了日 (sched_role columns). With no dated rows, falls back
    to a window around today so a fresh sheet still exports fillable week columns.
    Without this, export derives weeks only from existing EffortEntry rows, so an
    empty sheet has no week columns and 工数 can never be imported (要望: 工数Excel取込)."""
    start_col = next((c for c in columns if (c.config or {}).get("sched_role") == "start"), None)
    end_col = next((c for c in columns if (c.config or {}).get("sched_role") == "end"), None)
    dates: list[date] = []
    for r in rows:
        data = r.data or {}
        for col in (start_col, end_col):
            if col is None:
                continue
            d = _coerce_date_obj(data.get(str(col.id)))
            if d is not None:
                dates.append(d)
    today = date.today()
    if dates:
        lo, hi = min(dates), max(dates)
    else:
        lo, hi = today - timedelta(weeks=4), today + timedelta(weeks=12)
    w = week_start_of(lo, week_weekday)
    end = week_start_of(hi, week_weekday)
    out: list[date] = []
    while w <= end:
        out.append(w)
        w += timedelta(days=7)
    return out


def _sheet_weeks(db: Session, row_ids: list[int]) -> tuple[list[date], dict[tuple[int, date], EffortEntry]]:
    weeks: list[date] = []
    effort_map: dict[tuple[int, date], EffortEntry] = {}
    if not row_ids:
        return weeks, effort_map
    entries = list(
        db.execute(
            select(EffortEntry)
            .where(EffortEntry.row_id.in_(row_ids))
            .order_by(EffortEntry.week_start)
        ).scalars()
    )
    weeks = sorted({e.week_start for e in entries})
    for e in entries:
        effort_map[(e.row_id, e.week_start)] = e
    return weeks, effort_map


def _milestones_by_row(db: Session, row_ids: list[int]) -> dict[int, list[RowMilestone]]:
    out: dict[int, list[RowMilestone]] = {}
    if not row_ids:
        return out
    rows = db.execute(
        select(RowMilestone)
        .where(RowMilestone.row_id.in_(row_ids))
        .order_by(RowMilestone.row_id, RowMilestone.order, RowMilestone.boundary_date)
    ).scalars()
    for m in rows:
        out.setdefault(m.row_id, []).append(m)
    return out


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
@router.get("/{sheet_id}/export.xlsx")
def export_xlsx(
    sheet_id: int, user: User = Depends(current_user), db: Session = Depends(get_db)
) -> StreamingResponse:
    from openpyxl import Workbook

    sheet = get_sheet_for_user(db, sheet_id, user)
    ensure_schedule_columns(db, sheet)
    columns = list(
        db.execute(
            select(Column).where(Column.sheet_id == sheet.id).order_by(Column.order, Column.id)
        ).scalars()
    )
    rows = list(db.execute(select(Row).where(Row.sheet_id == sheet.id).order_by(Row.id)).scalars())
    row_ids = [r.id for r in rows]
    keyval_by_id = {r.id: (r.key_value or "") for r in rows}

    weeks: list[date] = []
    effort_map: dict[tuple[int, date], EffortEntry] = {}
    milestones_map: dict[int, list[RowMilestone]] = {}
    ms_cols: list[tuple[int, str, str, str]] = []
    if sheet.has_week_grid:
        weeks, effort_map = _sheet_weeks(db, row_ids)
        milestones_map = _milestones_by_row(db, row_ids)
        ms_cols = _template_milestone_cols(sheet)
        # Always export a fillable week range (spanning rows' 開始日〜完了日), unioned
        # with the weeks that already hold effort — so 工数 can be entered & imported
        # even on a sheet with no effort yet.
        week_weekday = org_week_start_weekday(db, user.org_id)
        weeks = sorted(set(weeks) | set(_span_weeks(rows, columns, week_weekday)))

    member_names = _members_by_id(db, user.org_id)

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet.name or "Sheet")[:31]

    # 進捗 / 先行タスク + one 予定/実績 pair per template ◇ — between attributes and weeks.
    extra_headers: list[str] = []
    if sheet.has_week_grid:
        extra_headers = [PROGRESS_HEADER, DEPS_HEADER]
        for _, _, ph, ah in ms_cols:
            extra_headers += [ph, ah]
    ws.append(
        ["ID"]
        + [c.name for c in columns]
        + extra_headers
        + [w.isoformat() for w in weeks]
    )

    today = date.today()
    for r in rows:
        data = r.data or {}
        attr: list = []
        for c in columns:
            v = data.get(str(c.id))
            if c.type in COMPUTED_COLUMN_TYPES:
                attr.append("")  # 計算列 — round-trips as blank
            elif c.type == "member":
                attr.append(member_names.get(str(v), "") if v not in (None, "") else "")
            elif c.type == "number":
                attr.append(_fmt_num(v) if v not in (None, "") else "")
            else:
                attr.append("" if v is None else str(v))

        extra_vals: list = []
        if sheet.has_week_grid:
            deps_keys = ",".join(
                keyval_by_id.get(pid, "") for pid in (r.depends_on or []) if keyval_by_id.get(pid)
            )
            extra_vals = [r.progress if r.progress is not None else "", deps_keys]
            # Match each template ◇ to the row's milestone of the same name (in order).
            by_name: dict[str, list[RowMilestone]] = {}
            for m in milestones_map.get(r.id, []):
                if m.kind == "milestone":
                    by_name.setdefault(m.name or "", []).append(m)
            for _, name, _ph, _ah in ms_cols:
                queue = by_name.get(name)
                m = queue.pop(0) if queue else None
                extra_vals += [
                    m.boundary_date.isoformat() if m and m.boundary_date else "",
                    m.actual_date.isoformat() if m and m.actual_date else "",
                ]

        week_vals: list = []
        for w in weeks:
            e = effort_map.get((r.id, w))
            if e is None:
                week_vals.append("")
                continue
            val = e.actual_hours if w < today else e.planned_hours
            week_vals.append(_fmt_num(val))
        ws.append([r.key_value or ""] + attr + extra_vals + week_vals)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"sheet_{sheet.id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --------------------------------------------------------------------------- #
# Import
# --------------------------------------------------------------------------- #
def _coerce_attr(column: Column, raw, members_by_name: dict[str, int]):
    """Convert an Excel cell to the value stored in row.data for this column."""
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        return None
    if column.type == "member":
        return members_by_name.get(str(raw).strip())
    if column.type == "number":
        try:
            f = float(raw)
            return int(f) if f == int(f) else f
        except (TypeError, ValueError):
            return None
    if column.type == "date":
        # 保存形は常に 'YYYY-MM-DD'（時刻は落とす）。読めない値はそのまま残す。
        return normalize_date_text(raw)
    # Alt+Enter line breaks inside a cell are kept — only the line ENDINGS are
    # normalised (要望: セル内の改行が1行になってしまう).
    # 日時セルを str() すると '2025-10-18 00:00:00' になるので _cell_text を通す
    # （0時ちょうどなら日付だけ）。数値の 3.0 → '3' もここで揃う。
    return _cell_text(raw)


def _parse_week_header(raw) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        try:
            return date.fromisoformat(raw.strip())
        except ValueError:
            return None
    return None


def looks_like_schedule(header: tuple) -> bool:
    """見出し行から「スケジュール形式（週の列がある）」かどうかを推測する。

    週の列は日付そのものが見出しになる。エクスポートしたブックなら 進捗(%) /
    先行タスク(ID) / ◇（予定・実績）も並ぶ。どれも無ければ、ただの表＝テーブル形式。
    一括取り込みは既定が「スケジュール」だったので、顧客リストのような表まで週グリッド
    付きで作られていた（要望: 取り込み時に形式を選べるように）。
    """
    for raw in header:
        if _parse_week_header(raw) is not None:
            return True
        name = str(raw).strip() if raw is not None else ""
        if name in (PROGRESS_HEADER, DEPS_HEADER):
            return True
        if name.endswith("（予定）") or name.endswith("（実績）"):
            return True
    return False


def _load_active_sheet(file: UploadFile):
    """Read an uploaded .xlsx and return (worksheet, header tuple, row iterator)."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file.file.read()), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excelファイルを読み込めませんでした（.xlsx 形式をご確認ください）",
        )
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="空のファイルです")
    return ws, header, rows_iter


@router.post("/{sheet_id}/import.xlsx")
def import_xlsx(
    sheet_id: int,
    file: UploadFile,
    sheet_name: str = Form(default=""),
    header_row: int = Form(default=0),
    last_row: int = Form(default=0),
    id_column: int = Form(default=0),
    match_mode: str = Form(default=""),
    columns: str = Form(default=""),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Import rows into an EXISTING sheet from an .xlsx.

    With no wizard fields this behaves exactly as before: the active worksheet,
    row 1 as the header, column A as the ID, every header matched to a column of
    the same name, upserted by ID. The 取り込みウィザード instead passes
    `sheet_name` / `header_row` / `last_row` / `id_column` / `match_mode` and
    `columns` — a JSON list of ``[{"index": 2, "name": "件名"}]`` mapping an Excel
    column to the sheet column (or reserved header) it should be written into.
    `last_row` is the last worksheet row to take (0 = 最後まで), for sheets that end
    in a 合計行 or notes. `match_mode` decides whether rows are matched at all
    (see :data:`MATCH_MODES`); omitted means the historic「ID列があれば照合」.
    """
    sheet = get_sheet_for_user(db, sheet_id, user)
    ensure_schedule_columns(db, sheet)
    mapping = _parse_selection(columns)
    if (
        mapping is None
        and not sheet_name
        and header_row <= 0
        and last_row <= 0
        and id_column == 0
        and not match_mode
    ):
        _ws, header, rows_iter = _load_active_sheet(file)
        return _import_rows(db, user, sheet, header, rows_iter)

    _wb, _ws, _grid, _hr, header, data_rows = xlsx.read_source(
        file, sheet_name, header_row, last_row
    )
    result = import_rows_with_mapping(
        db, user, sheet, header, data_rows, id_column, mapping, match_mode=match_mode
    )
    result.pop("selection", None)  # internal — only the 一括取り込み needs it
    return result


def import_rows_with_mapping(
    db: Session,
    user: User,
    sheet: Sheet,
    header: tuple,
    data_rows: list[tuple],
    id_column: int,
    mapping: list[dict] | None,
    commit: bool = True,
    match_mode: str = "",
) -> dict:
    """Import `data_rows` into an existing sheet through the wizard's column mapping.

    `mapping` is [{index, name}] — the Excel column and the sheet column (or
    reserved header) it goes into. None falls back to "every named column, matched
    by its own header". Shared with the 一括取り込み, which imports many worksheets
    inside one transaction (hence `commit`).

    `match_mode` decides what happens to the rows already in the sheet: keep them
    and match by ID ('id'), keep them and always add ('none'), or wipe them first
    ('replace' — 入れ替え, which is how a repeated load stays the same size instead
    of piling up copies).
    """
    mode = resolve_match_mode(match_mode, id_column)
    deleted = 0
    if mode == "replace":
        # 取り込む前に空にする。行・工数・◇・スナップショットが消え、採番も1に戻る
        # ので、同じファイルを何度入れ直しても中身は毎回まっさらから同じになる。
        from app.routers.sheets import clear_sheet_rows  # local import avoids a cycle

        deleted = clear_sheet_rows(db, sheet)
        db.flush()
    if mapping is None:
        mapping = [
            {"index": i, "name": _cell_text(h), "type": ""}
            for i, h in enumerate(header)
            if i != id_column and not _is_blank(h)
        ]
    kept = [
        {**it, "label": (it["name"] or _cell_text(xlsx.cell_at(header, it["index"]))).strip()}
        for it in mapping
        if it["index"] != id_column and 0 <= it["index"] < len(header)
    ]
    kept = [it for it in kept if it["label"]]
    out_header, out_rows = _synthesize(data_rows, kept, id_column)
    return {
        **_import_rows(
            db, user, sheet, out_header, iter(out_rows), commit=commit, match=mode == "id"
        ),
        # 入れ替えで消した行数。入れ替え以外では返さない（何も消していないので）。
        **({"deleted": deleted} if mode == "replace" else {}),
        # The mapping actually used — with `mapping=None` the defaults were filled
        # in here, and a preset has to remember those, not the None.
        "selection": _effective_selection(kept),
    }


def _effective_selection(kept: list[dict]) -> list[dict]:
    """The resolved picks in the `[{index, name, type, expr, lookup}]` shape a preset
    stores. 計算列（数式・参照）は中身まで残さないと、次回そこがただの値の列に戻る。"""
    return [
        {
            "index": it["index"],
            "name": it["label"],
            "type": it.get("type") or "",
            "expr": it.get("expr") or "",
            "lookup": it.get("lookup") or None,
        }
        for it in kept
    ]


@router.post("/{sheet_id}/import.xlsx/inspect")
def inspect_import_rows_xlsx(
    sheet_id: int,
    file: UploadFile,
    sheet_name: str = Form(default=""),
    header_row: int = Form(default=0),
    last_row: int = Form(default=0),
    tail_from: int = Form(default=0),
    id_column: int = Form(default=0),
    match_mode: str = Form(default=""),
    columns: str = Form(default=""),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Analyse an .xlsx against an EXISTING sheet without writing anything.

    Returns the worksheets, the guessed 見出し行, a head AND tail preview (the rows
    to cut off are at the bottom), the possible targets (this sheet's columns + the
    reserved schedule headers) with the proposed mapping per Excel column, how many
    rows would be 新規/更新, and how many values would not survive the conversion.
    Pass `columns` to re-check against the user's own mapping before committing.
    """
    sheet = get_sheet_for_user(db, sheet_id, user)
    ensure_schedule_columns(db, sheet)
    wb, ws, grid, hr, header, data_rows = xlsx.read_source(
        file, sheet_name, header_row, last_row
    )
    cols, targets = row_target_info(
        db, user, sheet, header, data_rows, id_column, _parse_selection(columns)
    )
    return {
        "worksheets": xlsx.worksheets_of(wb),
        "sheet_name": ws.title,
        "header_row": hr,
        "suggested_header_row": _auto_header_row(grid),
        "last_row": last_row,
        "sheet_last_row": len(grid),
        "id_column": id_column,
        "match_mode": resolve_match_mode(match_mode, id_column),
        "total_rows": len(data_rows),
        # Rows below the header with NO cut, so the UI can say how many the 最終行
        # setting is excluding.
        "available_rows": xlsx.data_row_total(grid, hr),
        "preview": xlsx.preview_of(grid),
        "tail_preview": xlsx.tail_preview_of(grid, hr, tail_from, last_row),
        "columns": cols,
        "targets": targets,
        **upsert_counts(db, sheet, data_rows, id_column, match_mode),
    }


def row_target_info(
    db: Session,
    user: User,
    sheet: Sheet,
    header: tuple,
    data_rows: list[tuple],
    id_column: int,
    chosen: list[dict] | None,
) -> tuple[list[dict], list[dict]]:
    """(per-Excel-column info, the targets it may be mapped onto) for an EXISTING
    sheet: the proposed/chosen target, the role and type it lands as, sample values
    and how many cells would not survive the conversion.

    Shared by the 既存シート wizard and the 一括取り込み dry-run, so both agree on
    what a saved mapping actually does.
    """
    sheet_columns = list(
        db.execute(
            select(Column).where(Column.sheet_id == sheet.id).order_by(Column.order, Column.id)
        ).scalars()
    )
    col_by_name = {c.name: c for c in sheet_columns}
    member_names = set(_members_by_name(db, user.org_id).keys())
    ms_labels: list[str] = []
    for _ti, _nm, ph, ah in _template_milestone_cols(sheet):
        ms_labels += [ph, ah]

    # What an Excel column can be mapped onto.
    targets: list[dict] = [
        {"key": c.name, "label": c.name, "type": c.type, "role": "attr"}
        for c in sheet_columns
        if c.type not in COMPUTED_COLUMN_TYPES
    ]
    if sheet.has_week_grid:
        targets.append({"key": PROGRESS_HEADER, "label": PROGRESS_HEADER, "type": "number", "role": "progress"})
        targets.append({"key": DEPS_HEADER, "label": DEPS_HEADER, "type": "text", "role": "deps"})
        targets += [
            {"key": lb, "label": lb, "type": "date", "role": "milestone"} for lb in ms_labels
        ]

    by_index = {it["index"]: it for it in (chosen or [])}

    cols: list[dict] = []
    for idx, raw_name in enumerate(header):
        nm = _cell_text(raw_name)
        values = _column_values(data_rows, idx)
        filled = [v for v in values if not _is_blank(v)]
        # Proposed target: the user's pick, else same-named column / reserved header.
        if idx in by_index:
            target = by_index[idx]["name"]
        elif idx == id_column or not nm:
            target = ""
        elif nm in col_by_name and col_by_name[nm].type not in COMPUTED_COLUMN_TYPES:
            target = nm
        elif sheet.has_week_grid and (
            nm in (PROGRESS_HEADER, DEPS_HEADER)
            or nm in ms_labels
            or _parse_week_header(raw_name) is not None
        ):
            target = nm
        else:
            target = ""

        # A chosen target the importer would NOT actually write has to stop being a
        # target here, or the dry run promises a column that silently goes nowhere.
        # Both cases show up after a sheet is edited post-import: the column was
        # renamed/deleted, or it was turned into a 計算列（参照(LOOKUP)／数式）—
        # `_import_rows` skips those by design, see `attr_at`).
        col = col_by_name.get(target)
        lost_target, lost_reason = "", ""
        if target and col is not None and col.type in COMPUTED_COLUMN_TYPES:
            lost_target, lost_reason, target, col = target, "computed", "", None
        elif (
            target
            and col is None
            and target not in (PROGRESS_HEADER, DEPS_HEADER)
            and target not in ms_labels
            and _parse_week_header(raw_name) is None
        ):
            lost_target, lost_reason, target = target, "missing", ""

        week = _parse_week_header(raw_name) if sheet.has_week_grid else None
        if col is not None:
            role, ctype = "attr", col.type
        elif target == PROGRESS_HEADER:
            role, ctype = "progress", "number"
        elif target == DEPS_HEADER:
            role, ctype = "deps", "text"
        elif target and target in ms_labels:
            role, ctype = "milestone", "date"
        elif target and week is not None:
            role, ctype = "week", "number"
        else:
            role, ctype = "", ""
        bad = _invalid_values(values, role if role in ("week", "progress") else "attr", ctype, member_names)
        cols.append(
            {
                "index": idx,
                "header": nm,
                "target": target,
                "role": role,
                "type": ctype,
                # The target that was asked for but cannot be written, and why
                # ('computed' | 'missing'). Empty when nothing was lost.
                "lost_target": lost_target,
                "lost_reason": lost_reason,
                # Week columns can only be kept or dropped — the date IS the target.
                "week_start": week.isoformat() if week is not None else None,
                "filled": len(filled),
                "samples": [_cell_text(v) for v in filled[:_SAMPLE_LIMIT]],
                "invalid": len(bad) if target else 0,
                "invalid_samples": bad[:_SAMPLE_LIMIT] if target else [],
            }
        )
    return cols, targets


def upsert_counts(
    db: Session,
    sheet: Sheet,
    data_rows: list[tuple],
    id_column: int,
    match_mode: str = "",
) -> dict:
    """How many rows would land on an existing task vs create a new one, plus the
    blank/duplicate ID counts the preview warns about.

    Only 'id' matching produces 更新 rows. 'none' adds every line; 'replace' does
    the same after emptying the sheet, so it also reports how many rows the run
    would delete (`deleted_rows`) — the number nobody wants to discover afterwards.
    """
    mode = resolve_match_mode(match_mode, id_column)
    existing_keys: set[str] = set()
    if mode == "id":
        existing_keys = {
            k
            for k in db.execute(select(Row.key_value).where(Row.sheet_id == sheet.id)).scalars()
            if k
        }
    ids = [_cell_text(xlsx.cell_at(r, id_column)) for r in data_rows] if id_column >= 0 else []
    seen: set[str] = set()
    updated = created = duplicate_ids = blank_ids = 0
    for k in ids:
        if not k:
            blank_ids += 1
            created += 1
            continue
        if k in seen:
            duplicate_ids += 1
        seen.add(k)
        if k in existing_keys:
            updated += 1
        else:
            created += 1
    if id_column < 0 or mode != "id":
        created = len(data_rows)
        updated = 0
    deleted = 0
    if mode == "replace":
        deleted = db.execute(
            select(func.count()).select_from(Row).where(Row.sheet_id == sheet.id)
        ).scalar_one()
    return {
        "new_rows": created,
        "updated_rows": updated,
        "deleted_rows": deleted,
        "blank_ids": blank_ids,
        "duplicate_ids": duplicate_ids,
    }


def _import_rows(
    db: Session,
    user: User,
    sheet: Sheet,
    header,
    rows_iter,
    commit: bool = True,
    match: bool = True,
) -> dict:
    """Import rows from an already-opened worksheet into `sheet`.

    `match=True` upserts by ID (key_value): a row whose ID already exists is
    updated. `match=False` never looks anything up — **every source line becomes
    its own row**, even when the ID column repeats (要望: 1列目が被るだけで1行に
    まとめられては困る). Blank IDs are auto-numbered from the sheet's 採番ルール, so
    a file with no ID column at all still gets a distinct internal id per row.

    `commit=False` leaves the transaction open so the 一括取り込み can write every
    worksheet of a workbook as one all-or-nothing operation."""
    columns = list(
        db.execute(
            select(Column).where(Column.sheet_id == sheet.id).order_by(Column.order, Column.id)
        ).scalars()
    )
    col_by_name = {c.name: c for c in columns}
    start_col, _ = sched_columns(db, sheet.id)
    start_col_id = str(start_col.id) if start_col else None
    template_items = _template_items(sheet)

    # Reserved schedule headers → template index + 予定/実績 role.
    ms_label_map: dict[str, tuple[str, int]] = {}
    for ti, _name, ph, ah in _template_milestone_cols(sheet):
        ms_label_map[ph] = ("pred", ti)
        ms_label_map[ah] = ("act", ti)

    # Map each header position (skipping col 0 = ID) to an attribute column, a
    # week_start date, or a reserved schedule column. Unknown/計算列 are ignored.
    attr_at: dict[int, Column] = {}
    week_at: dict[int, date] = {}
    ms_pred_at: dict[int, int] = {}  # header idx -> template index (予定)
    ms_act_at: dict[int, int] = {}   # header idx -> template index (実績)
    prog_idx: int | None = None
    deps_idx: int | None = None
    week_weekday = org_week_start_weekday(db, user.org_id)
    for idx, name in enumerate(header):
        if idx == 0 or name is None:
            continue
        nm = str(name).strip()
        col = col_by_name.get(nm)
        if col is not None:
            if col.type not in COMPUTED_COLUMN_TYPES:
                attr_at[idx] = col
            continue
        if sheet.has_week_grid:
            if nm == PROGRESS_HEADER:
                prog_idx = idx
                continue
            if nm == DEPS_HEADER:
                deps_idx = idx
                continue
            if nm in ms_label_map:
                role, ti = ms_label_map[nm]
                (ms_pred_at if role == "pred" else ms_act_at)[idx] = ti
                continue
            wk = _parse_week_header(name)
            if wk is not None:
                week_at[idx] = week_start_of(wk, week_weekday)

    members_by_name = _members_by_name(db, user.org_id)

    # Existing rows by key_value (first match wins, matching lookup semantics).
    # Only built when we are actually matching — with match=False nothing is looked
    # up, so two source lines sharing an ID can never collapse into one row.
    existing: dict[str, Row] = {}
    if match:
        for r in db.execute(
            select(Row).where(Row.sheet_id == sheet.id).order_by(Row.id)
        ).scalars():
            if r.key_value is not None:
                existing.setdefault(r.key_value, r)

    today = date.today()
    cur_week = current_week_start(week_weekday)
    created = updated = 0
    deps_to_resolve: list[tuple[Row, list[str]]] = []
    #: 取り込んだプルダウン列の値（列ID → 値の集合）。行を回し終えてから選択肢に足す。
    seen_dropdown_values: dict[int, set[str]] = {}

    for raw_row in rows_iter:
        if raw_row is None:
            continue
        id_cell = raw_row[0] if len(raw_row) > 0 else None
        key_value = None if id_cell is None else str(id_cell).strip()
        # Skip fully empty lines.
        if not any(v not in (None, "") for v in raw_row):
            continue

        row = existing.get(key_value) if (match and key_value) else None
        if row is None:
            if not key_value:
                key_value = _gen_key(db, sheet)
            row = Row(
                sheet_id=sheet.id,
                key_value=key_value,
                data={},
                version=1,
                created_by=user.id,
                updated_by=user.id,
            )
            db.add(row)
            db.flush()
            if match:
                existing[key_value] = row
            created += 1
        else:
            updated += 1

        data = dict(row.data or {})
        for idx, col in attr_at.items():
            if idx >= len(raw_row):
                continue
            value = _coerce_attr(col, raw_row[idx], members_by_name)
            data[str(col.id)] = value
            # プルダウン列は値を書くだけでなく、選択肢そのものも育てる。取り込みは
            # row.data に直接書くので、選択肢に無い値がそのまま入り、セルが「選択肢に
            # 未登録」だらけになっていた（要望: シート取込後にプルダウンがおかしい）。
            if col.type == "dropdown" and isinstance(value, str) and value:
                seen_dropdown_values.setdefault(col.id, set()).add(value)
        row.data = data
        row.version = (row.version or 1) + 1
        row.updated_by = user.id

        # 進捗(%)
        if prog_idx is not None and prog_idx < len(raw_row):
            pv = raw_row[prog_idx]
            if pv is not None and str(pv).strip() != "":
                try:
                    row.progress = max(0, min(100, int(float(pv))))
                    row.progress_week = cur_week
                except (TypeError, ValueError):
                    pass

        # 先行タスク(ID) — resolved to row ids after all rows exist.
        if deps_idx is not None and deps_idx < len(raw_row):
            dv = raw_row[deps_idx]
            if dv is not None and str(dv).strip() != "":
                deps_to_resolve.append(
                    (row, [s.strip() for s in str(dv).split(",") if s.strip()])
                )

        # ◇予定/◇実績 → rebuild this row's milestones from the template.
        predicted: dict[int, date] = {}
        actual: dict[int, date] = {}
        for idx, ti in ms_pred_at.items():
            if idx < len(raw_row):
                d = _coerce_date_obj(raw_row[idx])
                if d:
                    predicted[ti] = d
        for idx, ti in ms_act_at.items():
            if idx < len(raw_row):
                d = _coerce_date_obj(raw_row[idx])
                if d:
                    actual[ti] = d
        if predicted or actual:
            db.execute(delete(RowMilestone).where(RowMilestone.row_id == row.id))
            db.flush()
            start_iso = data.get(start_col_id) if start_col_id else None
            for kw in _reconstruct_milestones(template_items, predicted, actual, start_iso):
                db.add(RowMilestone(row_id=row.id, **kw))

        # Weekly effort upsert.
        for idx, wk in week_at.items():
            if idx >= len(raw_row):
                continue
            cell = raw_row[idx]
            if cell is None or (isinstance(cell, str) and cell.strip() == ""):
                continue
            try:
                hours = float(cell)
            except (TypeError, ValueError):
                continue
            entry = db.execute(
                select(EffortEntry).where(
                    EffortEntry.row_id == row.id, EffortEntry.week_start == wk
                )
            ).scalar_one_or_none()
            if entry is None:
                entry = EffortEntry(row_id=row.id, week_start=wk, version=1, updated_by=user.id)
                db.add(entry)
            if wk < today:
                entry.actual_hours = hours
            else:
                entry.planned_hours = hours
            entry.updated_by = user.id

    dropdown_notes = _grow_dropdown_options(columns, seen_dropdown_values)

    # Resolve 先行タスク (key_value → row id, first match in the sheet).
    if deps_to_resolve:
        db.flush()
        keymap: dict[str, int] = {}
        for r in db.execute(
            select(Row).where(Row.sheet_id == sheet.id).order_by(Row.id)
        ).scalars():
            if r.key_value is not None:
                keymap.setdefault(r.key_value, r.id)
        for r, keys in deps_to_resolve:
            r.depends_on = [keymap[k] for k in keys if k in keymap]

    if commit:
        db.commit()
    else:
        db.flush()
    return {"created": created, "updated": updated, "notes": dropdown_notes}


#: 取り込みで1列に自動追加してよい選択肢の上限。これを超える種類が入ってくる列は、
#: そもそもプルダウンではなく自由入力なので、勝手に何百件も足さずに知らせるだけにする。
_MAX_AUTO_OPTIONS = 60


def _grow_dropdown_options(
    columns: list[Column], seen: dict[int, set[str]]
) -> list[str]:
    """取り込んだ値をプルダウン列の選択肢に足す。返り値は画面に出す注意書き。

    足さないと、値は入っているのに選択肢に無い＝一覧で「選択肢に未登録」と出る状態に
    なる（要望: シート取込後にプルダウンがうまく追加できない）。ただし種類が
    `_MAX_AUTO_OPTIONS` を超える列まで機械的に足すと、住所のような自由記述が
    数百件の選択肢になって使い物にならないので、そこは足さずに理由を返す。
    """
    notes: list[str] = []
    by_id = {c.id: c for c in columns}
    for col_id, values in seen.items():
        col = by_id.get(col_id)
        if col is None or col.type != "dropdown":
            continue
        config = dict(col.config or {})
        options = list(config.get("options") or [])
        known = {o.get("value") for o in options if isinstance(o, dict)}
        fresh = sorted(v for v in values if v not in known)
        if not fresh:
            continue
        if len(known) + len(fresh) > _MAX_AUTO_OPTIONS:
            notes.append(
                f"「{col.name}」は取り込んだ値が {len(fresh)} 種類あり、選択肢には"
                f"追加していません（{_MAX_AUTO_OPTIONS} 種類が上限）。"
                "値はそのまま入っています。自由入力に変えるか、シート設定の"
                "「選択肢」から必要なものだけ追加してください。"
            )
            continue
        start = len(options)
        for i, v in enumerate(fresh):
            options.append(
                {
                    "id": uuid4().hex,
                    "value": v,
                    "color": _SWATCHES[(start + i) % len(_SWATCHES)],
                }
            )
        config["options"] = options
        col.config = config
        notes.append(f"「{col.name}」に選択肢を {len(fresh)} 件追加しました。")
    return notes


# --------------------------------------------------------------------------- #
# Import as a NEW sheet (シートごとExcelから取り込む)
# --------------------------------------------------------------------------- #
_DROPDOWN_MAX_OPTIONS = 20
# Palette reused from the dropdown options editor, so inferred lists look native.
_SWATCHES = [
    "#E3EFEA", "#EFEDE4", "#FAE6E0", "#E6F0DB",
    "#CBD9EE", "#F1DBAC", "#E7DDEA", "#DCE6EA",
]


_SAMPLE_LIMIT = xlsx.SAMPLE_LIMIT
_MAX_DROPDOWN_FORCED = 50   # cap when the user forces a dropdown by hand
_IMPORTABLE_TYPES = {"text", "number", "date", "dropdown", "member"}

# Generic workbook handling is shared with the 日報 importer (app.xlsx_import).
_is_blank = xlsx.is_blank
_cell_text = xlsx.cell_text
_auto_header_row = xlsx.auto_header_row
_split_grid = xlsx.split_grid
_column_values = xlsx.column_values
_looks_numeric = xlsx.looks_numeric


def _header_role(raw_name, has_week_grid: bool) -> str:
    """What a header means on a schedule sheet: attr / week / progress / deps /
    milestone. Only 'attr' (and an explicitly selected 'milestone') becomes a column;
    the rest are handled by the row importer under their reserved names."""
    if not has_week_grid:
        return "attr"
    nm = str(raw_name).strip()
    if nm == PROGRESS_HEADER:
        return "progress"
    if nm == DEPS_HEADER:
        return "deps"
    if _parse_week_header(raw_name) is not None:
        return "week"
    if nm.endswith("（予定）") or nm.endswith("（実績）"):
        return "milestone"
    return "attr"


def _infer_type(values: list, member_names: set[str]) -> tuple[str, dict]:
    """Guess a column type from its sample values. Returns (type, config)."""
    vals = [v for v in values if v is not None and str(v).strip() != ""]
    if not vals:
        return "text", {}
    strs = [_cell_text(v).strip() for v in vals]

    # 日付列の判定では「-」だけのセル（＝日付なし）を数に入れない。1個混ざっただけで
    # 列全体が自由入力に落ち、日時セルが '2025-10-18 00:00:00' で入っていた。
    dated = [v for v in vals if not is_date_placeholder(v)]
    if dated and all(_coerce_date_obj(v) is not None for v in dated):
        return "date", {}
    if all(_looks_numeric(v) for v in vals):
        return "number", {}
    # A cell filled in with Alt+Enter needs the 複数行入力 editor, or the grid's
    # single-line <input> flattens it on the first click. Never a pick-list either.
    if any("\n" in s for s in strs):
        return "text", {"multiline": True}
    # Every value naming an org member → 担当者 column.
    if member_names and all(s in member_names for s in strs):
        return "member", {}

    distinct = sorted(set(strs))
    # A short, repeating set of labels is a pick-list, not free text: at most 20
    # distinct values, and they must cover no more than 2/3 of the rows (so a
    # column of mostly-unique free text stays text).
    if 1 < len(distinct) <= _DROPDOWN_MAX_OPTIONS and len(distinct) * 3 <= len(strs) * 2:
        options = [
            {"id": uuid4().hex, "value": v, "color": _SWATCHES[i % len(_SWATCHES)]}
            for i, v in enumerate(distinct)
        ]
        return "dropdown", {"options": options}
    return "text", {}


def _synthesize(
    data_rows: list[tuple], kept: list[dict], id_column: int
) -> tuple[list, list[list]]:
    """A grid of exactly the chosen columns for `_import_rows`: ID first, then one
    column per pick under the header name it should be imported as."""
    out_header: list = ["ID"] + [it["label"] for it in kept]
    out_rows = [
        [xlsx.cell_at(r, id_column) if id_column >= 0 else None]
        + [xlsx.cell_at(r, it["index"]) for it in kept]
        for r in data_rows
    ]
    return out_header, out_rows


def _has_newline(values: list) -> bool:
    """Any sample value spans more than one line (Excel の Alt+Enter)."""
    return any("\n" in str(v) for v in values if not _is_blank(v))


def _dropdown_config(values: list) -> dict:
    """Options for a column the user explicitly typed as プルダウン."""
    distinct = sorted({str(v).strip() for v in values if not _is_blank(v)})[:_MAX_DROPDOWN_FORCED]
    return {
        "options": [
            {"id": uuid4().hex, "value": v, "color": _SWATCHES[i % len(_SWATCHES)]}
            for i, v in enumerate(distinct)
        ]
    }


def _parse_selection(raw: str) -> list[dict] | None:
    """`columns` form field → [{index, name, type}] (None when not supplied)."""
    items = xlsx.parse_json_field(raw, "取り込む列")
    if items is None:
        return None
    if not isinstance(items, list):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="取り込む列の指定を読み取れませんでした"
        )
    out: list[dict] = []
    for it in items:
        if not isinstance(it, dict):
            continue
        try:
            idx = int(it.get("index"))
        except (TypeError, ValueError):
            continue
        out.append(
            {
                "index": idx,
                "name": str(it.get("name") or "").strip(),
                "type": str(it.get("type") or "").strip(),
                # 数式列のとき、`[列名]` 形式の式。ウィザードが翻訳結果を送り返す。
                "expr": str(it.get("expr") or "").strip(),
                # 参照列のとき、XLOOKUP から起こした参照先（同じくウィザードが返す）。
                "lookup": _clean_lookup(it.get("lookup")),
            }
        )
    return out


#: 列名の上限（Column.name）。ここで切られた名前で列が作られる。
_NAME_MAX = 80


def formula_names_by_index(
    header: tuple, has_week_grid: bool, id_column: int, chosen: list[dict] | None
) -> dict[int, str]:
    """数式の中の A1 参照を置き換えるための「列位置 → **実際に作られる列名**」。

    ここが式の正しさそのもの。載せてよいのは **本当にシートの列になる見出しだけ** で、
    名前は **取り込み後に付く名前** でなければならない。ずれると、式は作られるのに
    参照先の列が無く、その列は全行 `#「◯◯」という列がありません` になる。

    実際に踏まれた食い違いは4つあった:

    * ウィザードで列名を打ち直した（`単価` → `価格`）— 見出しの名前で式を組んでいた
    * 取り込む列のチェックを外した — 列にならないのに参照していた
    * 見出しが80字を超える — 列名は切り詰められるのに、式は元の長さで参照していた
    * 見出しが重複している — 列は1本しか作られないのに、2つ目も参照できるつもりでいた

    週次工数 / 進捗 / 先行タスク / マイルストン は「列」にならないので載せない。

    **ID列だけは例外で、見出しの名前のまま載せる。** 列にはならないが、`=A2*2` や
    `[@品番]` が「どの列を指しているか」を突き止めるのに要るため。式に出力するときは
    `[ID]`（行のキー値）に翻訳される（:func:`xlsx_formula.translate_formula` の `emit`）
    ので、ここに載っている名前がそのまま式に出ることはない。
    """
    picked = {it["index"]: it for it in (chosen or [])}
    out: dict[int, str] = {}
    taken: set[str] = set()
    for idx, raw in enumerate(header):
        head = _cell_text(raw)
        if not head:
            continue
        if idx == id_column:
            out[idx] = head          # 特定用（式には出ない）
            continue
        if _header_role(raw, has_week_grid) != "attr":
            continue
        if chosen is not None and idx not in picked:
            continue          # 取り込まない列
        name = ((picked.get(idx, {}).get("name") or head).strip())[:_NAME_MAX].strip()
        # 先に出たほうが列になる（後から来た同名は作られない）ので、後勝ちにしない。
        if not name or name in taken:
            continue
        taken.add(name)
        out[idx] = name
    return out


def _clean_lookup(raw) -> dict | None:
    """ウィザードが返してきた参照先の指定を、使う4つだけに絞る。

    `local_index` は **Excel の列位置**。取り込み後の列IDはこの時点ではまだ存在しない
    （列を作るのはこれから）ので、位置で受け取って作成後に結び直す。
    """
    if not isinstance(raw, dict):
        return None
    try:
        sheet_id = int(raw.get("sheet_id"))
        local_index = int(raw.get("local_index"))
    except (TypeError, ValueError):
        return None
    match_id = str(raw.get("match_key_column_id") or "").strip()
    return_id = str(raw.get("return_column_id") or "").strip()
    if not match_id or not return_id:
        return None
    return {
        "sheet_id": sheet_id,
        "local_index": local_index,
        "match_key_column_id": match_id,
        "return_column_id": return_id,
    }


def _default_selection(
    header: tuple,
    data_rows: list[tuple],
    id_column: int,
    has_week_grid: bool,
    member_names: set[str],
    formula_rows: list[tuple[int, tuple]] | None = None,
    *,
    worksheet: str = "",
    tables: dict | None = None,
    resolve_lookup=None,
) -> list[dict]:
    """What the wizard proposes (and what an unattended import uses): every named
    column except the ID one, with its type inferred. ◇予定/◇実績 are left out —
    a brand-new sheet has no milestone template to hang them on.

    A column whose Excel cells are a translatable formula becomes a 数式列, and one
    holding a resolvable XLOOKUP/VLOOKUP becomes a 参照列 — keeping the calculation
    instead of the frozen values it happened to produce."""
    names_by_index = formula_names_by_index(header, has_week_grid, id_column, None)
    sel: list[dict] = []
    for idx, raw_name in enumerate(header):
        if idx == id_column or _is_blank(raw_name):
            continue
        nm = str(raw_name).strip()
        role = _header_role(raw_name, has_week_grid)
        if role == "milestone":
            continue
        ctype = ""
        expr = ""
        lookup = None
        if role == "attr":
            formula = _column_formula(
                formula_rows, idx, names_by_index,
                worksheet=worksheet, tables=tables, resolve_lookup=resolve_lookup,
                id_column=id_column,
            )
            resolved = (formula or {}).get("lookup")
            if formula and formula.get("expr"):
                ctype, expr = "formula", formula["expr"]
            elif resolved and resolved.get("ready"):
                ctype, lookup = "lookup", resolved
            else:
                ctype, _cfg = _infer_type(_column_values(data_rows, idx), member_names)
        sel.append({"index": idx, "name": nm, "type": ctype, "expr": expr, "lookup": lookup})
    return sel


def _invalid_values(values: list, role: str, ctype: str, member_names: set[str]) -> list[str]:
    """Cells that would be dropped or blanked on import, given the effective type."""
    bad: list[str] = []
    for v in values:
        if _is_blank(v):
            continue
        ok = True
        if role in ("week", "progress"):
            ok = _looks_numeric(v)
        elif role == "attr" or role == "milestone":
            if ctype == "date":
                ok = _coerce_date_obj(v) is not None
            elif ctype == "number":
                ok = _looks_numeric(v)
            elif ctype == "member":
                ok = str(v).strip() in member_names
        if not ok:
            bad.append(_cell_text(v))
    return bad


@router.post("/import.xlsx/inspect")
def inspect_import_xlsx(
    file: UploadFile,
    sheet_name: str = Form(default=""),
    header_row: int = Form(default=0),
    last_row: int = Form(default=0),
    tail_from: int = Form(default=0),
    id_column: int = Form(default=0),
    has_week_grid: bool = Form(default=True),
    columns: str = Form(default=""),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Analyse an .xlsx without writing anything — the 取り込みウィザード's data source.

    Returns the workbook's worksheets, a head AND tail preview of the chosen one
    (for picking the 見出し行 / ID列 / 最終行), and one entry per header column: its
    role on a schedule sheet, the guessed type, sample values and how many cells
    would NOT survive the conversion. Pass `columns` (the same JSON the import
    takes) to re-check the counts against the user's own choices before committing.
    """
    (
        wb, ws, grid, hr, header, data_rows, formula_rows, tables
    ) = xlsx.read_source_with_formulas(file, sheet_name, header_row, last_row)
    return {
        "worksheets": xlsx.worksheets_of(wb),
        "sheet_name": ws.title,
        "header_row": hr,
        "suggested_header_row": _auto_header_row(grid),
        "last_row": last_row,
        "sheet_last_row": len(grid),
        "id_column": id_column,
        "total_rows": len(data_rows),
        "available_rows": xlsx.data_row_total(grid, hr),
        "tail_preview": xlsx.tail_preview_of(grid, hr, tail_from, last_row),
        "preview": xlsx.preview_of(grid),
        "columns": new_sheet_column_info(
            db,
            user,
            header,
            data_rows,
            id_column,
            has_week_grid,
            _parse_selection(columns),
            formula_rows,
            worksheet=str(ws.title or ""),
            tables=tables,
            wb=wb,
        ),
        **id_column_counts(data_rows, id_column),
    }


def new_sheet_column_info(
    db: Session,
    user: User,
    header: tuple,
    data_rows: list[tuple],
    id_column: int,
    has_week_grid: bool,
    chosen: list[dict] | None,
    formula_rows: list[tuple[int, tuple]] | None = None,
    *,
    worksheet: str = "",
    tables: dict | None = None,
    wb=None,
) -> list[dict]:
    """Per-Excel-column analysis for a BRAND NEW sheet: the role the header plays,
    the guessed (or chosen) column type, sample values, inferred dropdown options
    and how many cells would not convert. Shared with the 一括取り込み dry-run.

    When `formula_rows` is given (the same rows read with `data_only=False`), each
    column also reports whether its cells are Excel FORMULAS and, if they translate,
    the equivalent `[列名]` expression — so the wizard can offer 数式列 instead of
    freezing today's numbers into text (要望: Excelの数式もいい感じに取り込む).

    A column whose formula is an XLOOKUP/VLOOKUP gets a `lookup` block instead: the
    sheet and columns it would point at, or the reason it cannot be wired up yet
    (要望: XLOOKUP使ってたらうまくLOOKUPを自動生成してほしい). 式の形が想定外で分解
    できなかった列には `lookup` が付かないが、`has_lookup` は必ず立つ — ウィザードは
    それを見て「参照先を手で選ぶ」入口を出す（:func:`_column_formula`）。"""
    member_names = set(_members_by_name(db, user.org_id).keys())
    by_index = {it["index"]: it for it in (chosen or [])}

    # 数式の中の A1 参照を置き換えるための対応表。**実際に作られる列名** で作る
    # （:func:`formula_names_by_index`）。ここがずれると、参照先の無い式ができる。
    names_by_index = formula_names_by_index(header, has_week_grid, id_column, chosen)
    resolve_lookup = (
        (
            lambda spec: _resolve_lookup(
                db, user, spec, wb=wb, names_by_index=names_by_index, worksheet=worksheet
            )
        )
        if wb is not None
        else None
    )

    cols: list[dict] = []
    for idx, raw_name in enumerate(header):
        nm = _cell_text(raw_name)
        role = _header_role(raw_name, has_week_grid) if nm else "attr"
        values = _column_values(data_rows, idx)
        picked = by_index.get(idx)
        if role == "attr" or (picked and role == "milestone"):
            inferred, cfg = _infer_type(values, member_names)
        else:
            inferred, cfg = "", {}

        formula = _column_formula(
            formula_rows, idx, names_by_index,
            worksheet=worksheet, tables=tables, resolve_lookup=resolve_lookup,
            id_column=id_column,
        )
        lookup = (formula or {}).get("lookup")
        # 数式が入っているのが **一部の行だけ** の列。残りは手で打った値なので、数式列に
        # すると計算値で上書きされて消える。列の型は1つしか持てない以上どちらかを捨てる
        # しかないが、**捨てるほうを既定にはしない** — 目に見えて入っている値が黙って
        # 消えるのが、いちばん気づけない壊れ方なので。
        if formula:
            formula["covers_all_rows"] = formula["cells"] >= len(data_rows)
            formula["replaced_values"] = max(0, len(data_rows) - formula["cells"])
        partial = bool(formula) and not formula["covers_all_rows"]

        # 翻訳できた列は、既定を「数式」／「参照」にする。ここが "いい感じ" の中身 —
        # 何もしなければ計算結果が焼き付いた文字列になり、元を直しても追従しない。
        if role == "attr" and not picked:
            if formula and formula.get("expr") and not partial:
                inferred = "formula"
            elif lookup and lookup.get("ready") and not partial:
                inferred = "lookup"

        ctype = (picked or {}).get("type") or inferred
        if ctype not in _IMPORTABLE_TYPES and ctype not in COMPUTED_COLUMN_TYPES:
            ctype = inferred
        filled = [v for v in values if not _is_blank(v)]
        # 計算列（数式・参照）は値を保存しないので、型が合わない件数の話は関係ない。
        bad = (
            []
            if ctype in COMPUTED_COLUMN_TYPES
            else _invalid_values(values, role, ctype, member_names)
        )
        selected = (
            idx in by_index
            if chosen is not None
            else (idx != id_column and bool(nm) and role != "milestone")
        )
        cols.append(
            {
                "index": idx,
                "header": nm,
                "role": role,
                "type": ctype,
                "selected": selected and idx != id_column,
                "filled": len(filled),
                "samples": [_cell_text(v) for v in filled[:_SAMPLE_LIMIT]],
                "options": [o["value"] for o in (cfg.get("options") or [])],
                "invalid": len(bad),
                "invalid_samples": bad[:_SAMPLE_LIMIT],
                **({"formula": formula} if formula else {}),
            }
        )
    return cols


def _column_formula(
    formula_rows: list[tuple[int, tuple]] | None,
    idx: int,
    names_by_index: dict[int, str],
    *,
    worksheet: str = "",
    tables: dict | None = None,
    resolve_lookup=None,
    id_column: int | None = None,
) -> dict | None:
    """この Excel 列が数式かどうかと、翻訳できるなら `[列名]` 式。数式が無ければ None。

    返す辞書はそのままウィザードに渡る:
      cells      … 数式セルの数
      expr       … 翻訳できた式（None なら翻訳不可）
      reason     … 翻訳できなかった理由
      sample     … 元の Excel 数式の例（画面に出して納得してもらうため）
      has_lookup … XLOOKUP/VLOOKUP が書いてあった列か（読み替えられたかとは別）
      lookup     … 参照列に読み替えられたときの中身（:func:`_resolve_lookup`）

    `has_lookup` と `lookup` を分けているのが肝。`lookup` は「キー列・照合列・取得列に
    分解できた」ときにしか入らないので、これを入口の条件にすると、式の書き方しだいで
    「参照先を選ぶ」ボタンが出たり出なかったりする（要望: XLOOKUP も参照が選べるものと
    選べないものがある。なぜ？）。画面が入口を出す条件は `has_lookup` — 分解できた分は
    初期値として使い、できなければ理由を見せて手で選んでもらう。

    `resolve_lookup` は「Excel の言葉で書かれた参照先」をアプリのシート／列に結びつける
    関数。DB を見る必要があるので呼び出し側から渡してもらう。
    """
    if not formula_rows:
        return None
    cells = [(row, xlsx.cell_at(r, idx)) for row, r in formula_rows]
    tr = xlsx_formula.translate_column(
        cells, names_by_index, worksheet=worksheet, tables=tables,
        # ID列は「シートの列」にならず行のID（key_value）になるので、それを指す式は
        # 列名ではなく `[ID]` に翻訳させる。
        id_index=id_column if (id_column is not None and id_column >= 0) else None,
    )
    if tr.formula_cells == 0:
        return None
    out = {
        "cells": tr.formula_cells,
        "expr": tr.expr,
        "reason": tr.reason,
        "sample": tr.sample,
        "has_lookup": tr.has_lookup,
    }
    if tr.lookup is not None and resolve_lookup is not None:
        resolved = resolve_lookup(tr.lookup)
        out["lookup"] = resolved
        # 参照としては読めたが結びつけられなかった、というのがいちばん多い。理由は
        # 「XLOOKUP に対応していない」ではなく「参照先が無い」なので、そちらを出す。
        if not resolved.get("ready"):
            out["reason"] = resolved.get("reason") or out["reason"]
    return out


def _resolve_lookup(
    db: Session,
    user: User,
    spec: xlsx_formula.LookupSpec,
    *,
    wb,
    names_by_index: dict[int, str],
    worksheet: str = "",
) -> dict:
    """XLOOKUP から起こした参照先を、このアプリのシートID・列IDに結びつける。

    Excel 側は「マスタ というワークシートの 品番 列」としか言っていない。アプリ側で
    それに当たるのは —

    1. その **ワークシートを取り込んだ先** のシート（取り込み設定に記録がある）。
       これがいちばん確か: 「このワークシートはこのシートに入れた」と本人が決めた記録。
    2. 同じ名前のシート。

    どちらも見つからなければ、参照列は作れない（作っても永久に空欄になる）。その場合は
    理由だけ返して、値としての取り込みに任せる。黙って空の参照列を作るより良い。
    """
    out: dict = {
        "local_index": spec.local_index,
        "local_column": names_by_index.get(spec.local_index) or "",
        "target_worksheet": spec.target_worksheet,
        "match_column": spec.match_column or "",
        "return_column": spec.return_column or "",
        "sheet_id": None,
        "sheet_name": "",
        "match_key_column_id": "",
        "return_column_id": "",
        "ready": False,
        "reason": None,
    }
    if not out["local_column"]:
        out["reason"] = "探すキーになる列が取り込み対象にありません"
        return out
    if worksheet and spec.target_worksheet == worksheet:
        # 自分自身を引いている式。参照列は「別のシートから引く」ものなので当てはまらない
        # （同じシート内で完結する話なら、そもそも参照ではなく普通の列でよい）。
        out["reason"] = "同じワークシートの中を参照しています（参照列にはできません）"
        return out

    preset = db.execute(
        select(ImportPreset).where(
            ImportPreset.org_id == user.org_id,
            ImportPreset.worksheet_name == spec.target_worksheet,
        )
    ).scalar_one_or_none()

    sheet = None
    if preset is not None and preset.target_sheet_id is not None:
        sheet = db.get(Sheet, preset.target_sheet_id)
        if sheet is not None and sheet.org_id != user.org_id:
            sheet = None
    if sheet is None:
        # 同じ名前のシートは作れてしまう。1本に絞れないなら **どれかを黙って選ばない** —
        # 画面には「マスタ」としか出ないので、どちらに繋がったのか利用者には区別が
        # つかないし、並び順の保証も無いので同じファイルを流し直すと別のシートを指す
        # 参照列ができうる。名前で決められないものは、手で選んでもらう。
        same_name = list(
            db.execute(
                select(Sheet)
                .where(Sheet.org_id == user.org_id, Sheet.name == spec.target_worksheet)
                .order_by(Sheet.id)
                .limit(2)
            ).scalars()
        )
        if len(same_name) > 1:
            out["reason"] = (
                f"「{spec.target_worksheet}」という名前のシートが複数あります"
                "（どれを参照するか選んでください）"
            )
            return out
        sheet = same_name[0] if same_name else None
    if sheet is None:
        out["reason"] = (
            f"参照先の「{spec.target_worksheet}」に当たるシートがまだありません"
            "（先にそのワークシートを取り込むと、参照列にできます）"
        )
        return out
    out["sheet_id"] = sheet.id
    out["sheet_name"] = sheet.name

    # 列番地（`マスタ!$C:$C`）で書かれていると列名が分からないので、参照先の見出し行を読む。
    target_headers = xlsx.worksheet_header_names(
        wb, spec.target_worksheet, preset.header_row if preset else 0
    )
    # `or` で書かないこと — A列（位置0）が「指定なし」に化ける。
    def header_at(pos: int | None) -> str:
        return target_headers.get(pos, "") if pos is not None else ""

    match_name = spec.match_column or header_at(spec.match_index)
    return_name = spec.return_column or header_at(spec.return_index)
    out["match_column"], out["return_column"] = match_name, return_name
    if not match_name or not return_name:
        out["reason"] = "参照先の列名が分かりませんでした（見出しのない列を指しています）"
        return out

    # 取り込みのとき ID列 にした列は、シートの列ではなく行のID（key_value）になっている
    # ので、名前で探しても見つからない。XLOOKUP がマスタを引くときに照合するのは
    # たいていその列なので、ここを拾えないと「参照先の列がありません」ばかりになる。
    #
    # ただし **どの列を ID にしたかは推測してはいけない**。以前は「取り込み設定に記録が
    # 無ければ先頭列」と決めつけていたが、単発の取り込みウィザードは設定を保存しないので、
    # これは例外ではなく常用経路だった。2列目を ID 列にして取り込んだマスタでは、Excel が
    # 「社内メモで照合」と書いているのに「品番の行IDと照合」する参照列を、警告も出さずに
    # `ready` で作ってしまう。
    #
    # そこで **値で確かめる**。行の key_value と、そのワークシート列の値が本当に重なって
    # いるなら、その列が ID になったと言い切れる。重ならなければ諦めて、手で選んでもらう。
    id_index = preset.id_column if preset is not None else 0
    id_header = target_headers.get(id_index, "") if id_index >= 0 else ""
    if id_header and not _column_became_the_id(
        db, sheet, wb, spec.target_worksheet, preset.header_row if preset else 0, id_index
    ):
        id_header = ""

    cols = list(
        db.execute(select(Column).where(Column.sheet_id == sheet.id)).scalars()
    )
    by_name = {(c.name or "").strip().casefold(): c for c in cols}

    def resolve(name: str) -> str:
        col = by_name.get(name.strip().casefold())
        if col is not None:
            return str(col.id)
        if id_header and name.strip().casefold() == id_header.strip().casefold():
            return LOOKUP_ID_KEY
        return ""

    out["match_key_column_id"] = resolve(match_name)
    out["return_column_id"] = resolve(return_name)
    missing = [
        nm
        for nm, got in ((match_name, out["match_key_column_id"]), (return_name, out["return_column_id"]))
        if not got
    ]
    if missing:
        out["reason"] = (
            f"「{sheet.name}」に " + "・".join(f"「{m}」" for m in missing) + " という列がありません"
        )
        return out
    out["ready"] = True
    return out


#: 「その列が行IDになった」と言い切るのに要る重なりの割合。
#: 取り込み後に行を足したり消したりしているのが普通なので、完全一致は求めない。
_ID_MATCH_RATIO = 0.5


def _column_became_the_id(
    db: Session, sheet: Sheet, wb, worksheet: str, header_row: int, index: int
) -> bool:
    """ワークシートの `index` 列が、`sheet` の行ID（key_value）になっているか。

    見出しの名前だけでは分からない（どの列を ID 列にして取り込んだかは記録が無いことも
    ある）ので、**実際の値が重なっているか** で判断する。取り込んだあとに行を足したり
    消したりしているのが普通なので、半分以上重なっていれば同じ列とみなす。

    重ならなければ False。呼び出し側は「参照先の列が見つからない」として、手で選ぶ道に
    落とす — 当てずっぽうで行IDに結びつけて、黙って違う列で照合するよりずっとよい。
    """
    values = xlsx.worksheet_column_values(wb, worksheet, header_row, index)
    if not values:
        return False
    keys = set(
        db.execute(
            select(Row.key_value).where(Row.sheet_id == sheet.id).limit(xlsx.SCAN_ROWS)
        ).scalars()
    )
    keys.discard(None)
    keys.discard("")
    if not keys:
        return False
    return len(keys & values) >= len(keys) * _ID_MATCH_RATIO


def id_column_counts(data_rows: list[tuple], id_column: int) -> dict:
    """Blank / duplicate ID counts — the warnings a NEW sheet's preview shows (every
    row is new there, so there is nothing to count as 更新)."""
    ids = [
        _cell_text(r[id_column]) if 0 <= id_column < len(r) else "" for r in data_rows
    ] if id_column >= 0 else []
    seen: set[str] = set()
    duplicate_ids = 0
    for k in ids:
        if not k:
            continue
        if k in seen:
            duplicate_ids += 1
        seen.add(k)
    return {"blank_ids": sum(1 for k in ids if not k), "duplicate_ids": duplicate_ids}


@router.post("/import.xlsx", status_code=status.HTTP_201_CREATED)
def import_new_sheet_xlsx(
    file: UploadFile,
    name: str = Form(default=""),
    has_week_grid: bool = Form(default=True),
    sheet_name: str = Form(default=""),
    header_row: int = Form(default=0),
    last_row: int = Form(default=0),
    id_column: int = Form(default=0),
    match_mode: str = Form(default=""),
    columns: str = Form(default=""),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Create a NEW sheet from an .xlsx and fill it (要望: シートもexcelから取り込める).

    `sheet_name` picks the worksheet (default: the active one), `header_row` the
    1-based 見出し行 (default: auto-detected), `last_row` the last worksheet row to
    take (0 → 最後まで; for sheets that end in a 合計行 or notes), `id_column` the
    0-based ID column (-1 → auto-numbered keys) and `columns` the JSON list of
    columns to take:
    ``[{"index": 1, "name": "件名", "type": "text"}]``. With `columns` omitted every
    named column is taken with its type guessed from the values (日付 / 数値 /
    メンバー / プルダウン / 自由入力).

    On a schedule sheet the reserved headers — 進捗(%), 先行タスク(ID) and ISO-date
    week columns — keep their normal meaning and do NOT become attribute columns;
    ◇予定/◇実績 are skipped unless explicitly selected (then they land as ordinary
    date columns, since a new sheet has no milestone template yet).
    """
    (
        wb,
        ws,
        _grid_rows,
        _hr,
        header,
        data_rows,
        formula_rows,
        tables,
    ) = xlsx.read_source_with_formulas(file, sheet_name, header_row, last_row)
    result = create_sheet_with_selection(
        db,
        user,
        name=name,
        has_week_grid=has_week_grid,
        worksheet_title=str(ws.title or ""),
        header=header,
        data_rows=data_rows,
        id_column=id_column,
        selection=_parse_selection(columns),
        formula_rows=formula_rows,
        match_mode=match_mode,
        tables=tables,
        wb=wb,
    )
    result.pop("selection", None)  # internal — only the 一括取り込み needs it
    return result


def create_sheet_with_selection(
    db: Session,
    user: User,
    *,
    name: str,
    has_week_grid: bool,
    worksheet_title: str,
    header: tuple,
    data_rows: list[tuple],
    id_column: int,
    selection: list[dict] | None,
    commit: bool = True,
    formula_rows: list[tuple[int, tuple]] | None = None,
    match_mode: str = "",
    tables: dict | None = None,
    wb=None,
) -> dict:
    """Create a sheet from an already-sliced worksheet and fill it.

    The body of `import_new_sheet_xlsx`, factored out so the 一括取り込み can create
    several sheets from one workbook inside a single transaction (`commit=False`).

    `formula_rows` (the same rows read with `data_only=False`) lets a column that
    holds an Excel formula become a 数式列 instead of the numbers it produced;
    `tables` + `wb` additionally let an XLOOKUP/VLOOKUP column become a 参照列.
    """
    if not any(not _is_blank(v) for v in header):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="見出し行が空です（見出し行の指定をご確認ください）"
        )

    member_names = set(_members_by_name(db, user.org_id).keys())
    if selection is None:
        names_by_index = formula_names_by_index(header, has_week_grid, id_column, None)
        selection = _default_selection(
            header, data_rows, id_column, has_week_grid, member_names, formula_rows,
            worksheet=worksheet_title,
            tables=tables,
            resolve_lookup=(
                (
                    lambda spec: _resolve_lookup(
                        db, user, spec, wb=wb, names_by_index=names_by_index,
                        worksheet=worksheet_title,
                    )
                )
                if wb is not None
                else None
            ),
        )
    selection = [it for it in selection if it["index"] != id_column and 0 <= it["index"] < len(header)]

    new_name = (name or "").strip() or worksheet_title.strip() or "取り込みシート"
    max_order = db.execute(
        select(func.coalesce(func.max(Sheet.order), -1)).where(Sheet.org_id == user.org_id)
    ).scalar_one()
    sheet = Sheet(
        org_id=user.org_id,
        name=new_name[:80],
        has_week_grid=has_week_grid,
        order=max_order + 1,
        numbering_rule={"prefix": "", "digits": 3, "next_seq": 1},
    )
    db.add(sheet)
    db.flush()
    # Schedule sheets get their 開始日/完了日 columns before inference, so those
    # headers bind to the real columns instead of creating duplicates.
    ensure_schedule_columns(db, sheet, commit=commit)

    existing_names = {
        c.name
        for c in db.execute(select(Column).where(Column.sheet_id == sheet.id)).scalars()
    }
    # Resolve each pick to its final column name once — the created columns and the
    # synthesized header the row importer sees must line up exactly.
    kept = [
        {**it, "label": ((it["name"] or _cell_text(header[it["index"]])).strip())[:80].strip()}
        for it in selection
    ]
    kept = [it for it in kept if it["label"]]

    created_cols = 0
    order = len(existing_names)
    # 参照列は「同じシートのキー列」を指すので、その列が出来てからでないと設定を書けない。
    # 場所（order）だけ先に取っておいて、本体は下の2周目で作る。
    pending_lookups: list[tuple[dict, int]] = []
    # この取り込みで **本当に出来る列名**。式の検算に使う。
    # `expr` は画面から送り返されてくる値なので、こちら側でも必ず確かめる — 画面で列名を
    # 打ち直したあと、古い名前のままの式が届く、という食い違いが実際に起きていた。
    will_exist = set(existing_names)
    for it in kept:
        role = _header_role(header[it["index"]], has_week_grid)
        if role not in ("week", "progress", "deps"):
            will_exist.add(it["label"])
    if id_column >= 0:
        will_exist.add(xlsx_formula.ID_REF)   # `[ID]` は行のキー値

    for it in kept:
        idx, nm = it["index"], it["label"]
        role = _header_role(header[idx], has_week_grid)
        # 週次工数 / 進捗 / 先行タスク are imported by name, not as attribute columns.
        if role in ("week", "progress", "deps") or nm in existing_names:
            continue
        values = _column_values(data_rows, idx)
        # 計算列（数式・参照）は「取り込める型」の一覧には入らない（値を保存しないので）が、
        # 選択としては正当。中身が空なら普通の推定に落とす — 式や参照先の無い計算列は
        # 永久に空欄になるだけなので。
        #
        # 式が「作られない列」を指しているときも同じ扱い（値として取り込む）。読めない式で
        # 列を作ると、その列は全行エラー表示になり、元の Excel にあった値まで失う。
        wants_formula = (
            it["type"] == "formula"
            and bool(it.get("expr"))
            and xlsx_formula.expr_refs(it["expr"]) <= will_exist
        )
        wants_lookup = it["type"] == "lookup" and bool(it.get("lookup"))
        if wants_lookup:
            existing_names.add(nm)
            pending_lookups.append((it, order))
            order += 1
            created_cols += 1
            continue
        ctype = "formula" if wants_formula else (it["type"] if it["type"] in _IMPORTABLE_TYPES else "")
        if wants_formula:
            config = {"expr": it["expr"]}
        elif not ctype:
            ctype, config = _infer_type(values, member_names)
        elif ctype == "dropdown":
            config = _dropdown_config(values)
        elif ctype == "text" and _has_newline(values):
            config = {"multiline": True}   # 複数行入力（セル内改行を保てるように）
        else:
            config = {}
        db.add(
            Column(
                sheet_id=sheet.id,
                name=nm,
                type=ctype,
                order=order,
                is_key=False,
                config=config,
            )
        )
        existing_names.add(nm)
        order += 1
        created_cols += 1
    db.flush()
    _create_lookup_columns(
        db, user, sheet, pending_lookups, kept, id_column, data_rows, member_names
    )

    out_header, out_rows = _synthesize(data_rows, kept, id_column)
    # 新しいシートなので照合する相手はいない。効くのは「同じファイル内で ID が重なった
    # 行」で、'id' 以外なら 1行ずつ別の行として入る（要望: 勝手にまとめない）。
    result = _import_rows(
        db,
        user,
        sheet,
        out_header,
        iter(out_rows),
        commit=commit,
        match=resolve_match_mode(match_mode, id_column) == "id",
    )
    return {
        "sheet_id": sheet.id,
        "name": sheet.name,
        "columns": created_cols,
        # With `selection=None` the picks were inferred here; a preset must store
        # what was inferred so the next run repeats it against the sheet just made.
        "selection": _effective_selection(kept),
        **result,
    }


def _points_at_a_real_column(col_id: str, target_cols: set[str]) -> bool:
    """参照先の指定が、そのシートに本当にある列か。

    `__id__` は列ではなく行のキー値なので、列の一覧には無くて正しい。
    """
    return col_id == LOOKUP_ID_KEY or col_id in target_cols


def _create_lookup_columns(
    db: Session,
    user: User,
    sheet: Sheet,
    pending: list[tuple[dict, int]],
    kept: list[dict],
    id_column: int,
    data_rows: list[tuple],
    member_names: set[str],
) -> None:
    """XLOOKUP から起こした参照列を作る（他の列を作り終えた **あと** に呼ぶ）。

    参照列の設定は「このシートのどの列をキーにするか」を列IDで持つので、キーになる列が
    実在してからでないと書けない。Excel の列位置 → さっき作った列、と結び直すのがここ。

    参照先が解決できないときは、参照列にせず **値の列として作る**。中身の分からない
    参照列を置いても永久に空欄になるだけで、元の Excel にあった値まで消えてしまうので。

    参照先はウィザードで手で選べる（名前がずれていて自動では結びつかないとき）ので、
    ここに届く指定は「画面が言っているだけ」の値。**自分のグループのシートか／その
    シートに本当にある列か** を、必ずこちらで確かめてから書く。
    """
    if not pending:
        return
    by_name = {
        (c.name or ""): c
        for c in db.execute(select(Column).where(Column.sheet_id == sheet.id)).scalars()
    }
    label_by_index = {it["index"]: it["label"] for it in kept}
    sheet_ids = set(
        db.execute(select(Sheet.id).where(Sheet.org_id == user.org_id)).scalars()
    )
    # 参照先シートごとの、実在する列ID。手で選んだ指定を確かめるのに使う。
    wanted = {(it.get("lookup") or {}).get("sheet_id") for it, _o in pending}
    cols_by_sheet: dict[int, set[str]] = {sid: set() for sid in wanted & sheet_ids}
    if cols_by_sheet:
        for col in db.execute(
            select(Column).where(Column.sheet_id.in_(list(cols_by_sheet)))
        ).scalars():
            cols_by_sheet[col.sheet_id].add(str(col.id))

    for it, order in pending:
        cfg = it.get("lookup") or {}
        local_index = cfg.get("local_index")
        # キー列は、ID列そのもの（行の key_value）か、いま作った普通の列のどちらか。
        if local_index == id_column:
            local_key = LOOKUP_ID_KEY
        else:
            local_col = by_name.get(label_by_index.get(local_index, ""))
            local_key = str(local_col.id) if local_col is not None else ""

        target_cols = cols_by_sheet.get(cfg.get("sheet_id"))
        target_ok = target_cols is not None and all(
            _points_at_a_real_column(cfg.get(k, ""), target_cols)
            for k in ("match_key_column_id", "return_column_id")
        )
        if local_key and target_ok:
            ctype = "lookup"
            config: dict = {
                "target_sheet_id": cfg["sheet_id"],
                "local_key_column_id": local_key,
                "match_key_column_id": cfg["match_key_column_id"],
                "return_column_id": cfg["return_column_id"],
            }
        else:
            ctype, config = _infer_type(
                _column_values(data_rows, it["index"]), member_names
            )
        db.add(
            Column(
                sheet_id=sheet.id,
                name=it["label"],
                type=ctype,
                order=order,
                is_key=False,
                config=config,
            )
        )
    db.flush()


def _gen_key(db: Session, sheet: Sheet) -> str:
    """Next key_value from the sheet numbering rule (advances next_seq)."""
    rule = dict(sheet.numbering_rule or {})
    prefix = str(rule.get("prefix", ""))
    digits = int(rule.get("digits", 3))
    next_seq = int(rule.get("next_seq", 1))
    key = f"{prefix}{next_seq:0{digits}d}"
    rule["next_seq"] = next_seq + 1
    sheet.numbering_rule = rule
    return key
